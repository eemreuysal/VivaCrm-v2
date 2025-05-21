"""
Category export functionality.
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd

from products.models import Category
from products.forms.category import CategorySearchForm


@login_required
def export_categories(request):
    """
    Export categories to Excel or CSV format.
    """
    # Get format from URL parameters
    export_format = request.GET.get('format', 'xlsx').lower()
    
    # Get filtered queryset
    queryset = Category.objects.all()
    
    # Add annotations for counts
    queryset = queryset.select_related('parent').annotate(
        product_count=Count('products'),
        active_product_count=Count('products', filter=Q(products__is_active=True))
    )
    
    # Apply filters from request
    form = CategorySearchForm(request.GET)
    if form.is_valid():
        queryset = form.filter_queryset(queryset)
    
    # Create Excel file
    if export_format == 'csv':
        return export_categories_csv(queryset)
    else:
        return export_categories_excel(queryset)


def export_categories_excel(queryset):
    """Export categories to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Kategoriler"
    
    # Headers
    headers = [
        'ID', 'Kategori Adı', 'Üst Kategori', 'Açıklama', 
        'Durum', 'Ürün Sayısı', 'Aktif Ürün Sayısı', 'Oluşturulma Tarihi'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Write data
    for row_num, category in enumerate(queryset, 2):
        values = [
            category.id,
            category.name,
            category.parent.name if category.parent else 'Ana Kategori',
            category.description or '',
            'Aktif' if category.is_active else 'Pasif',
            category.product_count,
            category.active_product_count,
            category.created_at.strftime('%Y-%m-%d %H:%M') if category.created_at else ''
        ]
        
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
    
    # Add category hierarchy sheet
    ws_hierarchy = wb.create_sheet(title="Kategori Hiyerarşisi")
    
    # Hierarchy headers
    hierarchy_headers = ['Seviye', 'Kategori Adı', 'Ürün Sayısı', 'Yol']
    
    for col_num, header in enumerate(hierarchy_headers, 1):
        cell = ws_hierarchy.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write hierarchy
    row_num = 2
    for category in queryset.filter(parent__isnull=True).order_by('name'):
        # Root category
        values = [
            1,
            category.name,
            category.product_count,
            category.name
        ]
        
        for col_num, value in enumerate(values, 1):
            cell = ws_hierarchy.cell(row=row_num, column=col_num, value=value)
        
        row_num += 1
        
        # Child categories
        for child in queryset.filter(parent=category).order_by('name'):
            values = [
                2,
                child.name,
                child.product_count,
                f"{category.name} > {child.name}"
            ]
            
            for col_num, value in enumerate(values, 1):
                cell = ws_hierarchy.cell(row=row_num, column=col_num, value=value)
                if col_num == 2:  # Indent child categories
                    cell.alignment = Alignment(indent=2)
            
            row_num += 1
    
    # Adjust column widths
    for ws_obj in [ws, ws_hierarchy]:
        for column in ws_obj.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_obj.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=kategoriler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


def export_categories_csv(queryset):
    """Export categories to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=kategoriler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    df_list = []
    
    for category in queryset:
        df_list.append({
            'ID': category.id,
            'Kategori Adı': category.name,
            'Üst Kategori': category.parent.name if category.parent else 'Ana Kategori',
            'Açıklama': category.description or '',
            'Durum': 'Aktif' if category.is_active else 'Pasif',
            'Ürün Sayısı': category.product_count,
            'Aktif Ürün Sayısı': category.active_product_count,
            'Oluşturulma Tarihi': category.created_at.strftime('%Y-%m-%d %H:%M') if category.created_at else '',
            'Kategori Yolu': get_category_path(category)
        })
    
    df = pd.DataFrame(df_list)
    df.to_csv(response, index=False, encoding='utf-8-sig')
    
    return response


def get_category_path(category):
    """Get the full path of a category."""
    path = [category.name]
    parent = category.parent
    
    while parent:
        path.insert(0, parent.name)
        parent = parent.parent
    
    return ' > '.join(path)