"""
Excel import/export functionality for the Products app.
"""
import pandas as pd
import re
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from django.http import HttpResponse
from decimal import Decimal
import requests
from urllib.parse import urlparse
from django.core.files.base import ContentFile

from .models import (
    Product, Category, ProductFamily, ProductAttribute, 
    ProductAttributeValue, ProductImage, StockMovement
)


def generate_product_import_template():
    """
    Generate an Excel template for importing products.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    
    # Define column headers based on the requested format
    headers = [
        "SKU", "URUNISMI", "BARKOD", "ASIN", "KATEGORI", "FIYAT", 
        "URUNMALIYETI", "KARGOMALIYET", "KOMISYON", "URUNAILESI", 
        "RENK", "BOYUT", "GORSELURL"
    ]
    
    # Add column headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # Add sample data rows
    sample_data = [
        [
            "WU-FN33-YUYN", "Viva Maison Adult Muslin Towels Set for Bathroom, 4-Layer Pure Turkish Cotton Muslin Fabric 2-Piece Bath and Hand Towel (Sage Green, 2-Piece Hand & Bath Towel Set)", 
            "8684612428258", "B0DXQGSFCQ", "Muslin Towel", "44.65", "1,00", "1,00", "1,00",
            "Viva Maison Adult Muslin Towels Set for Bathroom, 4-Layer Pure Turkish Cotton Muslin Fabric 2-Piece Bath and Hand Towel", 
            "Sage Green", "2-Piece Hand & Bath Towel Set", "https://m.media-amazon.com/images/I/81eg1n1NDyL._AC_SL1500_.jpg"
        ],
        [
            "9I-QLHR-2VVF", "Viva Maison Adult Muslin Towels Set for Bathroom, 4-Layer Pure Turkish Cotton Muslin Fabric 2-Piece Bath and Hand Towel (Black, 2-Piece Hand & Bath Towel Set)", 
            "8684612428241", "B0DXQDY19P", "Muslin Towel", "44.65", "1,00", "1,00", "1,00",
            "Viva Maison Adult Muslin Towels Set for Bathroom, 4-Layer Pure Turkish Cotton Muslin Fabric 2-Piece Bath and Hand Towel", 
            "Black", "2-Piece Hand & Bath Towel Set", "https://m.media-amazon.com/images/I/71u95DZRTSL.__AC_SX300_SY300_QL70_FMwebp_.jpg"
        ]
    ]
    
    # Add sample data to the sheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # For better readability of the data
            if col_idx in [6, 7, 8, 9]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Add instructions in a new sheet
    ws2 = wb.create_sheet(title="Talimatlar")
    
    ws2.cell(row=1, column=1, value="Ürün İçe Aktarma Talimatları").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Gerekli Alanlar:").font = Font(bold=True)
    ws2.cell(row=4, column=1, value="- SKU: Ürün Kodu (zorunlu)")
    ws2.cell(row=5, column=1, value="- URUNISMI: Ürün adı (zorunlu)")
    ws2.cell(row=6, column=1, value="- FIYAT: Ürün fiyatı (zorunlu, nokta ile ayrılmış, örn: 44.65)")
    
    ws2.cell(row=8, column=1, value="İsteğe Bağlı Alanlar:").font = Font(bold=True)
    ws2.cell(row=9, column=1, value="- BARKOD: Ürün barkod numarası")
    ws2.cell(row=10, column=1, value="- ASIN: Amazon ürün kimliği")
    ws2.cell(row=11, column=1, value="- KATEGORI: Ürün kategorisi")
    ws2.cell(row=12, column=1, value="- URUNMALIYETI: Ürün maliyeti (virgül ile ayrılmış, örn: 1,00)")
    ws2.cell(row=13, column=1, value="- KARGOMALIYET: Kargo maliyeti (virgül ile ayrılmış, örn: 1,00)")
    ws2.cell(row=14, column=1, value="- KOMISYON: Komisyon maliyeti (virgül ile ayrılmış, örn: 1,00)")
    ws2.cell(row=15, column=1, value="- URUNAILESI: Ürün ailesi")
    ws2.cell(row=16, column=1, value="- RENK: Ürün rengi")
    ws2.cell(row=17, column=1, value="- BOYUT: Ürün boyutu/ölçüsü")
    ws2.cell(row=18, column=1, value="- GORSELURL: Ürün görseli URL adresi")
    
    ws2.cell(row=20, column=1, value="Önemli Notlar:").font = Font(bold=True)
    ws2.cell(row=21, column=1, value="1. İlk satır başlık satırıdır, değiştirmeyiniz")
    ws2.cell(row=22, column=1, value="2. SKU, URUNISMI ve FIYAT alanları zorunludur, diğer alanlar isteğe bağlıdır")
    ws2.cell(row=23, column=1, value="3. FIYAT alanında nokta (.), diğer sayısal değerlerde virgül (,) kullanabilirsiniz")
    ws2.cell(row=24, column=1, value="4. Aynı SKU'ya sahip bir ürün zaten varsa, bilgileri güncellenecektir")
    
    # Adjust column widths in the instructions sheet
    ws2.column_dimensions['A'].width = 80
    
    # Auto-adjust column widths in the main sheet
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = ws.column_dimensions[chr(64 + col)]
        for cell in ws[chr(64 + col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid excessive widths
        column.width = adjusted_width
    
    # Save to response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=urun_excel_sablonu.xlsx'
    
    return response


def import_products_excel(file_obj, update_existing=True):
    """
    Import products from an Excel file with the new format.
    
    Args:
        file_obj: The uploaded Excel file
        update_existing: Whether to update existing products by SKU
        
    Returns:
        dict: Result statistics and error messages
    """
    # Initialize the results dictionary
    results = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'error_count': 0,
        'error_rows': []
    }
    
    try:
        # Read Excel file
        df = pd.read_excel(file_obj)
        
        # Update total count
        results['total'] = len(df)
        
        # Verify the expected columns
        required_columns = ["SKU", "URUNISMI", "FIYAT"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Gerekli sütunlar eksik: {', '.join(missing_columns)}")
        
        # Process each row
        for index, row in df.iterrows():
            try:
                row_num = index + 2  # +2 because Excel is 1-indexed and we have a header row
                
                # Validate required fields
                sku = row.get('SKU')
                if pd.isna(sku) or not sku:
                    raise ValueError("SKU zorunludur")
                
                product_name = row.get('URUNISMI')
                if pd.isna(product_name) or not product_name:
                    raise ValueError("Ürün ismi zorunludur")
                
                # Parse price value
                price_str = str(row.get('FIYAT'))
                try:
                    # First try with period as decimal separator
                    price_float = float(price_str)
                except ValueError:
                    try:
                        # Then try with comma as decimal separator
                        price_float = float(price_str.replace(',', '.'))
                    except ValueError:
                        raise ValueError(f"Geçersiz fiyat: {price_str}")
                
                # En fazla 2 ondalık basamağa yuvarla (model kısıtlaması)
                from decimal import Decimal, ROUND_HALF_UP
                price = Decimal(str(price_float)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                # Check if product already exists
                existing_product = None
                try:
                    existing_product = Product.objects.get(code=sku)
                except Product.DoesNotExist:
                    pass
                
                # Create a unique slug - Her SATIR için tamamen benzersiz SLUG oluştur
                from django.utils.text import slugify
                import uuid
                
                # Ürün adı yerine SKU ile birlikte benzersiz slug oluştur
                # UUID ekleyerek kesinlikle benzersiz olmasını sağla
                unique_id = str(uuid.uuid4())[:8]
                base_slug = f"{slugify(sku)}-{unique_id}"
                
                # Eğer slug zaten varsa, timestamp ekle
                while Product.objects.filter(slug=base_slug).exists():
                    base_slug = f"{slugify(sku)}-{unique_id}-{int(timezone.now().timestamp())}"

                # Simple description 
                barcode = row.get('BARKOD', '') if not pd.isna(row.get('BARKOD', '')) else ""
                description = f"İçe aktarılan ürün - {timezone.now().strftime('%Y-%m-%d')}"
                    
                # Create or update product - ULTRA SİMPLİFİYE EDİLMİŞ - SADECE KRİTİK ALANLAR
                # Sinyalleri devre dışı bırak - böylece StockMovement oluşturma işlemi çalışmaz
                from django.db.models.signals import post_save
                from django.db import models
                
                # Geçici olarak sinyalleri devre dışı bırak
                post_save_handlers = models.signals.post_save.receivers
                models.signals.post_save.receivers = []
                
                try:
                    if existing_product and update_existing:
                        # Update existing product - önce temel alanlar
                        existing_product.name = product_name
                        existing_product.price = price
                        existing_product.barcode = barcode
                        
                        # Kategori ve ürün ailesi işlemleri
                        category_name = row.get('KATEGORI', '') if not pd.isna(row.get('KATEGORI', '')) else None
                        family_name = row.get('URUNAILESI', '') if not pd.isna(row.get('URUNAILESI', '')) else None
                        
                        # Kategori güncelleme
                        if category_name:
                            try:
                                category = Category.objects.filter(name__iexact=category_name).first()
                                if not category:
                                    category_slug = slugify(category_name)
                                    if Category.objects.filter(slug=category_slug).exists():
                                        category_slug = f"{category_slug}-{int(timezone.now().timestamp())}"
                                    category = Category.objects.create(
                                        name=category_name, 
                                        slug=category_slug
                                    )
                                existing_product.category = category
                            except Exception as e:
                                print(f"Kategori güncelleme hatası: {str(e)}")
                        
                        # Ürün ailesi güncelleme
                        if family_name:
                            try:
                                family = ProductFamily.objects.filter(name__iexact=family_name).first()
                                if not family:
                                    family_slug = slugify(family_name)
                                    if ProductFamily.objects.filter(slug=family_slug).exists():
                                        family_slug = f"{family_slug}-{int(timezone.now().timestamp())}"
                                    # Ürün Ailesi oluştururken hata çıkmasına karşı detaylı logla
                                    try:
                                        family = ProductFamily.objects.create(
                                            name=family_name,
                                            slug=family_slug
                                        )
                                        print(f"Ürün ailesi başarıyla oluşturuldu: {family_name}")
                                    except Exception as family_create_error:
                                        print(f"Ürün ailesi oluşturma detaylı hata: {str(family_create_error)}")
                                        print(f"Ürün ailesi adı: {family_name}")
                                        print(f"Ürün ailesi adı uzunluğu: {len(family_name)}")
                                        print(f"Ürün ailesi slug: {family_slug}")
                                        print(f"Ürün ailesi slug uzunluğu: {len(family_slug)}")
                                        # Ürün ailesi oluşturma başarısız olursa None olarak ayarla
                                        family = None
                                existing_product.family = family
                            except Exception as e:
                                print(f"Ürün ailesi güncelleme hatası: {str(e)}")
                        
                        # Bilgileri kaydet
                        existing_product.save()
                        
                        # Renk ve boyut bilgilerini işle
                        color = row.get('RENK', '') if not pd.isna(row.get('RENK', '')) else None
                        size = row.get('BOYUT', '') if not pd.isna(row.get('BOYUT', '')) else None
                        image_url = row.get('GORSELURL', '') if not pd.isna(row.get('GORSELURL', '')) else None
                        
                        # Renk özelliğini güncelle
                        if color:
                            try:
                                # Renk özelliğini bul veya oluştur
                                attr_renk, _ = ProductAttribute.objects.get_or_create(
                                    name='Renk',
                                    defaults={'slug': 'renk'}
                                )
                                # Ürünün renk değerini güncelle
                                attr_val, created = ProductAttributeValue.objects.get_or_create(
                                    product=existing_product,
                                    attribute=attr_renk,
                                    defaults={'value': color}
                                )
                                # Mevcut değeri güncelle
                                if not created and attr_val.value != color:
                                    attr_val.value = color
                                    attr_val.save()
                            except Exception as e:
                                print(f"Renk güncelleme hatası: {str(e)}")
                        
                        # Boyut özelliğini güncelle
                        if size:
                            try:
                                # Boyut özelliğini bul veya oluştur
                                attr_boyut, _ = ProductAttribute.objects.get_or_create(
                                    name='Boyut',
                                    defaults={'slug': 'boyut'}
                                )
                                # Ürünün boyut değerini güncelle
                                attr_val, created = ProductAttributeValue.objects.get_or_create(
                                    product=existing_product,
                                    attribute=attr_boyut,
                                    defaults={'value': size}
                                )
                                # Mevcut değeri güncelle
                                if not created and attr_val.value != size:
                                    attr_val.value = size
                                    attr_val.save()
                            except Exception as e:
                                print(f"Boyut güncelleme hatası: {str(e)}")
                        
                        # Ürün görseli ekle (ürünün görseli yoksa)
                        if image_url and not ProductImage.objects.filter(product=existing_product).exists():
                            try:
                                # URL'i doğrula
                                result = urlparse(image_url)
                                if all([result.scheme, result.netloc]):
                                    # Görseli indir
                                    response = requests.get(image_url, timeout=10)
                                    if response.status_code == 200:
                                        # Dosya adını URL'den al
                                        filename = image_url.split('/')[-1]
                                        if not filename or '.' not in filename:
                                            filename = f"{sku}.jpg"
                                        
                                        # Görsel kaydını oluştur
                                        content = ContentFile(response.content)
                                        image = ProductImage(
                                            product=existing_product,
                                            is_primary=True  # İlk görsel ana görsel olsun
                                        )
                                        image.image.save(filename, content, save=True)
                            except Exception as e:
                                print(f"Görsel ekleme hatası: {str(e)}")
                        
                        results['updated'] += 1
                    else:
                        # Kategori ve Ürün Ailesi Bilgilerini İşle
                        category_name = row.get('KATEGORI', '') if not pd.isna(row.get('KATEGORI', '')) else None
                        family_name = row.get('URUNAILESI', '') if not pd.isna(row.get('URUNAILESI', '')) else None
                        
                        # Kategori oluştur veya bul
                        category = None
                        if category_name:
                            try:
                                # İlk olarak mevcut kategoriyi bul
                                category = Category.objects.filter(name__iexact=category_name).first()
                                # Yoksa oluştur
                                if not category:
                                    category_slug = slugify(category_name)
                                    # Eğer slug zaten varsa, benzersiz yap
                                    if Category.objects.filter(slug=category_slug).exists():
                                        category_slug = f"{category_slug}-{int(timezone.now().timestamp())}"
                                    category = Category.objects.create(
                                        name=category_name,
                                        slug=category_slug
                                    )
                            except Exception as e:
                                print(f"Kategori oluşturma hatası: {str(e)}")
                        
                        # Ürün Ailesi oluştur veya bul
                        family = None
                        if family_name:
                            try:
                                # İlk olarak mevcut ürün ailesini bul
                                family = ProductFamily.objects.filter(name__iexact=family_name).first()
                                # Yoksa oluştur
                                if not family:
                                    family_slug = slugify(family_name)
                                    # Eğer slug zaten varsa, benzersiz yap
                                    if ProductFamily.objects.filter(slug=family_slug).exists():
                                        family_slug = f"{family_slug}-{int(timezone.now().timestamp())}"
                                    # Ürün Ailesi oluştururken hata çıkmasına karşı detaylı logla
                                    try:
                                        family = ProductFamily.objects.create(
                                            name=family_name,
                                            slug=family_slug
                                        )
                                        print(f"Ürün ailesi başarıyla oluşturuldu: {family_name}")
                                    except Exception as family_create_error:
                                        print(f"Ürün ailesi oluşturma detaylı hata: {str(family_create_error)}")
                                        print(f"Ürün ailesi adı: {family_name}")
                                        print(f"Ürün ailesi adı uzunluğu: {len(family_name)}")
                                        print(f"Ürün ailesi slug: {family_slug}")
                                        print(f"Ürün ailesi slug uzunluğu: {len(family_slug)}")
                                        # Ürün ailesi oluşturma başarısız olursa None olarak ayarla
                                        family = None
                            except Exception as e:
                                print(f"Ürün ailesi oluşturma hatası: {str(e)}")
                        
                        # Create new product - SADECE KRİTİK ZORUNLU ALANLAR
                        product = Product.objects.create(
                            code=sku,           # Zorunlu
                            name=product_name,  # Zorunlu
                            slug=base_slug,     # Zorunlu - benzersiz
                            sku=sku,            # Opsiyonel ama dolduruyoruz
                            barcode=barcode,    # Opsiyonel 
                            description=description,  # Opsiyonel
                            price=price,        # Zorunlu
                            stock=0,            # Varsayılan değer
                            tax_rate=18,        # Varsayılan değer
                            is_active=True,     # Varsayılan değer
                            category=category,  # Opsiyonel - null olabilir
                            family=family       # Opsiyonel - null olabilir
                        )
                        
                        # Renk, Boyut ve Görsel bilgilerini işle
                        color = row.get('RENK', '') if not pd.isna(row.get('RENK', '')) else None
                        size = row.get('BOYUT', '') if not pd.isna(row.get('BOYUT', '')) else None
                        image_url = row.get('GORSELURL', '') if not pd.isna(row.get('GORSELURL', '')) else None
                        
                        # Renk özelliğini ekle
                        if color:
                            try:
                                # Renk özelliğini bul veya oluştur
                                attr_renk, _ = ProductAttribute.objects.get_or_create(
                                    name='Renk',
                                    defaults={'slug': 'renk'}
                                )
                                # Ürüne renk değerini ekle
                                ProductAttributeValue.objects.create(
                                    product=product,
                                    attribute=attr_renk,
                                    value=color
                                )
                            except Exception as e:
                                print(f"Renk özelliği ekleme hatası: {str(e)}")
                        
                        # Boyut özelliğini ekle
                        if size:
                            try:
                                # Boyut özelliğini bul veya oluştur
                                attr_boyut, _ = ProductAttribute.objects.get_or_create(
                                    name='Boyut',
                                    defaults={'slug': 'boyut'}
                                )
                                # Ürüne boyut değerini ekle
                                ProductAttributeValue.objects.create(
                                    product=product,
                                    attribute=attr_boyut,
                                    value=size
                                )
                            except Exception as e:
                                print(f"Boyut özelliği ekleme hatası: {str(e)}")
                        
                        # Ürün görseli ekle
                        if image_url:
                            try:
                                # URL'i doğrula
                                result = urlparse(image_url)
                                if all([result.scheme, result.netloc]):
                                    # Görseli indir
                                    response = requests.get(image_url, timeout=10)
                                    if response.status_code == 200:
                                        # Dosya adını URL'den al
                                        filename = image_url.split('/')[-1]
                                        if not filename or '.' not in filename:
                                            filename = f"{sku}.jpg"
                                        
                                        # Görsel kaydını oluştur
                                        content = ContentFile(response.content)
                                        image = ProductImage(
                                            product=product,
                                            is_primary=True  # İlk görsel ana görsel olsun
                                        )
                                        image.image.save(filename, content, save=True)
                            except Exception as e:
                                print(f"Görsel ekleme hatası: {str(e)}")
                        
                        results['created'] += 1
                finally:
                    # Sinyalleri geri yükle
                    models.signals.post_save.receivers = post_save_handlers
                
            except Exception as e:
                results['error_count'] += 1
                results['error_rows'].append({
                    'row': row_num,
                    'error': str(e)
                })
        
        return results
        
    except Exception as e:
        # Handle file reading/format errors
        raise Exception(f"Excel dosyası işlenirken hata oluştu: {str(e)}")


class StockAdjustmentExcelImporter:
    """Stok düzeltme Excel import sınıfı - core.views_import tarafından kullanılıyor"""
    
    @staticmethod
    def import_data(file, user=None):
        """Excel dosyasından stok düzeltme verilerini içe aktar"""
        from .excel import import_stock_adjustments_excel
        return import_stock_adjustments_excel(file)
    
    @staticmethod
    def get_template():
        """Stok düzeltme import şablonunu oluştur"""
        from .excel import generate_product_template
        return generate_product_template()