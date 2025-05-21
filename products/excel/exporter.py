"""
Ürün Excel export işlemleri.
Core base class'ları extend eder.
"""
from typing import List, Dict, Any
from django.db import models
from django.utils import timezone
import logging
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from core.excel.base import BaseExcelExporter
from products.models import Product

logger = logging.getLogger(__name__)


class ProductExcelExporter(BaseExcelExporter):
    """
    Ürün-spesifik Excel exporter.
    Formatlanmış ve optimize edilmiş export işlemleri.
    """
    
    def __init__(self):
        super().__init__(Product)
        
    def get_export_fields(self) -> List[str]:
        """Export edilecek field'ların listesi"""
        return [
            'id',
            'code',
            'name',
            'category__name',
            'family__name',
            'price',
            'sale_price',
            'sale_end_date',
            'cost',
            'tax_rate',
            'current_stock',
            'threshold_stock',
            'sku',
            'barcode',
            'is_physical',
            'weight',
            'dimensions',
            'status',
            'created_at',
            'updated_at'
        ]
    
    def get_field_headers(self) -> Dict[str, str]:
        """Field isimlerini Excel başlıklarına map'le"""
        return {
            'id': 'ID',
            'code': 'Ürün Kodu',
            'name': 'Ürün Adı',
            'category__name': 'Kategori',
            'family__name': 'Ürün Ailesi',
            'price': 'Fiyat',
            'sale_price': 'İndirimli Fiyat',
            'sale_end_date': 'İndirim Bitiş Tarihi',
            'cost': 'Maliyet',
            'tax_rate': 'KDV Oranı (%)',
            'current_stock': 'Stok',
            'threshold_stock': 'Minimum Stok',
            'sku': 'SKU',
            'barcode': 'Barkod',
            'is_physical': 'Fiziksel Ürün',
            'weight': 'Ağırlık (kg)',
            'dimensions': 'Boyutlar',
            'status': 'Durum',
            'created_at': 'Oluşturma Tarihi',
            'updated_at': 'Güncelleme Tarihi'
        }
    
    def _optimize_queryset(self, queryset, fields: List[str]):
        """QuerySet'i optimize et"""
        # Related field'ları bul
        related_fields = []
        
        for field in fields:
            if '__' in field:
                related_field = field.split('__')[0]
                if related_field not in related_fields:
                    related_fields.append(related_field)
        
        # select_related uygula
        if related_fields:
            queryset = queryset.select_related(*related_fields)
        
        # Sadece gerekli field'ları seç
        return queryset.only(*fields)
    
    def _format_value(self, value):
        """Değeri Excel için formatla"""
        if value is None:
            return ''
        
        # Boolean değerleri Türkçeleştir
        if isinstance(value, bool):
            return 'Evet' if value else 'Hayır'
        
        # Status değerlerini Türkçeleştir
        if isinstance(value, str) and value in dict(Product.STATUS_CHOICES):
            status_dict = dict(Product.STATUS_CHOICES)
            return status_dict.get(value, value)
        
        # Tarih formatı
        if hasattr(value, 'strftime'):
            return value.strftime('%d.%m.%Y %H:%M')
        
        return value
    
    def apply_formatting(self, worksheet):
        """Excel formatlaması uygula"""
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Başlık satırını formatla
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Kolon genişliklerini ayarla
        column_widths = {
            'A': 8,   # ID
            'B': 15,  # Ürün Kodu
            'C': 35,  # Ürün Adı
            'D': 20,  # Kategori
            'E': 20,  # Ürün Ailesi
            'F': 12,  # Fiyat
            'G': 12,  # İndirimli Fiyat
            'H': 18,  # İndirim Bitiş
            'I': 12,  # Maliyet
            'J': 10,  # KDV Oranı
            'K': 10,  # Stok
            'L': 12,  # Min Stok
            'M': 15,  # SKU
            'N': 15,  # Barkod
            'O': 12,  # Fiziksel Ürün
            'P': 10,  # Ağırlık
            'Q': 15,  # Boyutlar
            'R': 15,  # Durum
            'S': 18,  # Oluşturma
            'T': 18   # Güncelleme
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # Para formatı
        money_format = '#,##0.00 "₺"'
        money_columns = ['F', 'G', 'I']  # Fiyat, İndirimli Fiyat, Maliyet
        
        for col in money_columns:
            for cell in worksheet[col][1:]:  # Başlık hariç
                if cell.value:
                    cell.number_format = money_format
        
        # Yüzde formatı
        percentage_columns = ['J']  # KDV Oranı
        
        for col in percentage_columns:
            for cell in worksheet[col][1:]:  # Başlık hariç
                if cell.value:
                    cell.number_format = '0.00"%"'
        
        # Satır yüksekliği
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20
    
    def export_with_stock_info(self, queryset, **options):
        """Stok bilgileri ile detaylı export"""
        # Stok durumuna göre renkli export
        from io import BytesIO
        import pandas as pd
        from openpyxl import load_workbook
        
        # Normal export yap
        excel_data = self.export_data(queryset, **options)
        
        # Excel'i yeniden aç ve stok durumuna göre renklendir
        workbook = load_workbook(BytesIO(excel_data))
        worksheet = workbook.active
        
        # Stok durumuna göre renklendirme
        low_stock_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        out_of_stock_fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
        
        # Stok ve Minimum Stok kolonlarını bul
        stock_col = None
        threshold_col = None
        
        for cell in worksheet[1]:
            if cell.value == 'Stok':
                stock_col = cell.column
            elif cell.value == 'Minimum Stok':
                threshold_col = cell.column
        
        if stock_col and threshold_col:
            for row in worksheet.iter_rows(min_row=2):
                stock_cell = row[stock_col - 1]
                threshold_cell = row[threshold_col - 1]
                
                if stock_cell.value is not None and threshold_cell.value is not None:
                    stock = float(stock_cell.value)
                    threshold = float(threshold_cell.value)
                    
                    if stock == 0:
                        # Stok yok - kırmızı
                        for cell in row:
                            cell.fill = out_of_stock_fill
                    elif stock <= threshold:
                        # Düşük stok - sarı
                        for cell in row:
                            cell.fill = low_stock_fill
        
        # Kaydet ve döndür
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_price_list(self, queryset, **options):
        """Fiyat listesi export'u (sadeleştirilmiş)"""
        # Sadece fiyat bilgileri
        self.get_export_fields = lambda: [
            'code',
            'name',
            'category__name',
            'price',
            'sale_price',
            'tax_rate',
            'status'
        ]
        
        self.get_field_headers = lambda: {
            'code': 'Ürün Kodu',
            'name': 'Ürün Adı',
            'category__name': 'Kategori',
            'price': 'Fiyat',
            'sale_price': 'İndirimli Fiyat',
            'tax_rate': 'KDV Oranı (%)',
            'status': 'Durum'
        }
        
        return self.export_data(queryset, **options)
    
    def export_inventory_report(self, queryset, **options):
        """Envanter raporu export'u"""
        # Envanter bilgileri
        self.get_export_fields = lambda: [
            'code',
            'name',
            'category__name',
            'current_stock',
            'threshold_stock',
            'cost',
            'price',
            'barcode',
            'status'
        ]
        
        self.get_field_headers = lambda: {
            'code': 'Ürün Kodu',
            'name': 'Ürün Adı',
            'category__name': 'Kategori',
            'current_stock': 'Mevcut Stok',
            'threshold_stock': 'Minimum Stok',
            'cost': 'Birim Maliyet',
            'price': 'Satış Fiyatı',
            'barcode': 'Barkod',
            'status': 'Durum'
        }
        
        # Stok değeri hesaplaması için özel alan ekle
        excel_data = self.export_data(queryset, **options)
        
        # DataFrame'e dönüştür ve stok değeri ekle
        import pandas as pd
        from io import BytesIO
        
        df = pd.read_excel(BytesIO(excel_data))
        
        # Stok değeri hesapla
        df['Stok Değeri'] = df['Mevcut Stok'] * df['Birim Maliyet']
        df['Tahmini Satış Değeri'] = df['Mevcut Stok'] * df['Satış Fiyatı']
        
        # Toplam satırı ekle
        totals = pd.DataFrame({
            'Ürün Kodu': ['TOPLAM'],
            'Ürün Adı': [''],
            'Kategori': [''],
            'Mevcut Stok': [df['Mevcut Stok'].sum()],
            'Minimum Stok': [''],
            'Birim Maliyet': [''],
            'Satış Fiyatı': [''],
            'Barkod': [''],
            'Durum': [''],
            'Stok Değeri': [df['Stok Değeri'].sum()],
            'Tahmini Satış Değeri': [df['Tahmini Satış Değeri'].sum()]
        })
        
        df = pd.concat([df, totals], ignore_index=True)
        
        # Excel'e dönüştür
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Envanter Raporu')
            
            # Formatting
            worksheet = writer.sheets['Envanter Raporu']
            self.apply_formatting(worksheet)
            
            # Toplam satırını vurgula
            last_row = worksheet.max_row
            total_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            total_font = Font(bold=True)
            
            for cell in worksheet[last_row]:
                cell.fill = total_fill
                cell.font = total_font
        
        output.seek(0)
        return output.getvalue()