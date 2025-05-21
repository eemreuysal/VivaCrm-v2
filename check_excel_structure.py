#!/usr/bin/env python
"""Check Excel file structure and content."""
import pandas as pd
import os

excel_path = '/Users/emreuysal/Downloads/siparis_excel_sablonu.xlsx'

print(f"Checking Excel file: {excel_path}")
print(f"File exists: {os.path.exists(excel_path)}")
print(f"File size: {os.path.getsize(excel_path) / 1024 / 1024:.2f} MB")

try:
    # Read first few rows to understand structure
    df_head = pd.read_excel(excel_path, nrows=5)
    print(f"\nColumns: {list(df_head.columns)}")
    print(f"Shape: {df_head.shape}")
    print("\nFirst 5 rows:")
    print(df_head)
    
    # Check total rows
    df_info = pd.read_excel(excel_path)
    print(f"\nTotal rows: {len(df_info)}")
    print(f"Total columns: {len(df_info.columns)}")
    
    # Memory usage
    memory_usage = df_info.memory_usage(deep=True).sum() / 1024 / 1024
    print(f"DataFrame memory usage: {memory_usage:.2f} MB")
    
except Exception as e:
    print(f"Error reading Excel: {str(e)}")
    
    # Try reading with openpyxl
    try:
        print("\nTrying with openpyxl...")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        print(f"Sheet name: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Read headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"Headers: {headers}")
        
    except Exception as e2:
        print(f"Error with openpyxl: {str(e2)}")