"""
Customer Excel işlemlerini yöneten facade sınıfı.
Facade Pattern: Karmaşık alt sistemlere basit bir arayüz sağlar.
"""
from typing import Optional, Dict, Any, Union
from pathlib import Path
import logging
from django.http import HttpResponse
from django.utils import timezone
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

from ..services import CustomerImportService
from core.excel.exceptions import ExcelError

logger = logging.getLogger(__name__)


class CustomerExcelManager:
    """
    Müşteri Excel işlemlerini yöneten ana sınıf.
    Import, export ve diğer Excel işlemleri için tek bir arayüz sağlar.
    """
    
    # Customer export için tanımlar
    CUSTOMER_EXPORT_FIELDS = [
        'id', 'name', 'type', 'company_name', 'tax_office', 'tax_number',
        'email', 'phone', 'website', 'owner__username', 'is_active',
        'created_at', 'notes'
    ]
    
    CUSTOMER_EXPORT_HEADERS = {
        'id': 'ID',
        'name': 'Müşteri Adı',
        'type': 'Tip',
        'company_name': 'Şirket Adı',
        'tax_office': 'Vergi Dairesi',
        'tax_number': 'Vergi/TC Kimlik No',
        'email': 'Email',
        'phone': 'Telefon',
        'website': 'Website',
        'owner__username': 'Sorumlu',
        'is_active': 'Aktif',
        'created_at': 'Kayıt Tarihi',
        'notes': 'Notlar'
    }
    
    # Address export için tanımlar
    ADDRESS_EXPORT_FIELDS = [
        'id', 'customer__name', 'title', 'type', 'address_line1', 'address_line2',
        'city', 'state', 'postal_code', 'country', 'is_default'
    ]
    
    ADDRESS_EXPORT_HEADERS = {
        'id': 'ID',
        'customer__name': 'Müşteri',
        'title': 'Adres Başlığı',
        'type': 'Adres Tipi',
        'address_line1': 'Adres Satırı 1',
        'address_line2': 'Adres Satırı 2',
        'city': 'Şehir',
        'state': 'İlçe',
        'postal_code': 'Posta Kodu',
        'country': 'Ülke',
        'is_default': 'Varsayılan Adres'
    }
    
    def import_customers_excel(
        self,
        file_obj,
        update_existing: bool = True
    ) -> Dict[str, Any]:
        """
        Excel dosyasından müşteri verilerini içe aktar.
        
        Args:
            file_obj: Excel dosyası
            update_existing: Mevcut müşterileri güncelle
            
        Returns:
            dict: İçe aktarma istatistikleri
        """
        try:
            return CustomerImportService.import_customers_excel(
                file_obj,
                update_existing=update_existing
            )
                
        except ExcelError as e:
            logger.error(f"Excel import hatası: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Beklenmeyen import hatası: {str(e)}")
            raise ExcelError(f"Import işlemi başarısız: {str(e)}")
            
    def import_addresses_excel(
        self,
        file_obj,
        update_existing: bool = True
    ) -> Dict[str, Any]:
        """
        Excel dosyasından adres verilerini içe aktar.
        
        Args:
            file_obj: Excel dosyası
            update_existing: Mevcut adresleri güncelle
            
        Returns:
            dict: İçe aktarma istatistikleri
        """
        try:
            return CustomerImportService.import_addresses_excel(
                file_obj,
                update_existing=update_existing
            )
                
        except ExcelError as e:
            logger.error(f"Excel import hatası: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Beklenmeyen import hatası: {str(e)}")
            raise ExcelError(f"Import işlemi başarısız: {str(e)}")
    
    def export_customers_excel(self, queryset) -> bytes:
        """
        Müşterileri Excel formatında dışa aktar.
        
        Args:
            queryset: Dışa aktarılacak müşteri queryset'i
            
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            return self._create_excel_file(
                queryset=queryset,
                fields=self.CUSTOMER_EXPORT_FIELDS,
                headers=self.CUSTOMER_EXPORT_HEADERS,
                sheet_name='Müşteriler'
            )
                
        except Exception as e:
            logger.error(f"Excel export hatası: {str(e)}")
            raise ExcelError(f"Export işlemi başarısız: {str(e)}")
    
    def export_customers_csv(self, queryset) -> bytes:
        """
        Müşterileri CSV formatında dışa aktar.
        
        Args:
            queryset: Dışa aktarılacak müşteri queryset'i
            
        Returns:
            bytes: CSV dosya içeriği
        """
        try:
            # Veriyi hazırla
            data = []
            for obj in queryset:
                row = {}
                for field in self.CUSTOMER_EXPORT_FIELDS:
                    if field == 'type':
                        row[field] = obj.get_type_display() if hasattr(obj, 'get_type_display') else getattr(obj, field)
                    elif '__' in field:
                        parts = field.split('__')
                        val = obj
                        for part in parts:
                            if val is not None:
                                val = getattr(val, part) if hasattr(val, part) else None
                        row[field] = val
                    else:
                        row[field] = getattr(obj, field) if hasattr(obj, field) else None
                data.append(row)
            
            # DataFrame oluştur
            df = pd.DataFrame(data)
            
            # Sütun başlıkları uygula
            df = df.rename(columns=self.CUSTOMER_EXPORT_HEADERS)
            
            # CSV dosyasına dönüştür
            buffer = BytesIO()
            df.to_csv(buffer, index=False, encoding='utf-8-sig')
            buffer.seek(0)
            return buffer.getvalue()
                
        except Exception as e:
            logger.error(f"CSV export hatası: {str(e)}")
            raise ExcelError(f"CSV export işlemi başarısız: {str(e)}")
    
    def export_addresses_excel(self, queryset) -> bytes:
        """
        Adresleri Excel formatında dışa aktar.
        
        Args:
            queryset: Dışa aktarılacak adres queryset'i
            
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            return self._create_excel_file(
                queryset=queryset,
                fields=self.ADDRESS_EXPORT_FIELDS,
                headers=self.ADDRESS_EXPORT_HEADERS,
                sheet_name='Adresler'
            )
                
        except Exception as e:
            logger.error(f"Excel export hatası: {str(e)}")
            raise ExcelError(f"Export işlemi başarısız: {str(e)}")
    
    def generate_customer_import_template(self) -> bytes:
        """
        Müşteri içe aktarma için Excel şablonu oluştur.
        
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Müşteriler"
            
            # Sütun başlıklarını tanımla
            headers = {
                'name': 'Müşteri Adı *',
                'type': 'Tip (individual/corporate) *',
                'company_name': 'Şirket Adı',
                'tax_office': 'Vergi Dairesi',
                'tax_number': 'Vergi/TC Kimlik No',
                'email': 'Email *',
                'phone': 'Telefon *',
                'website': 'Website',
                'notes': 'Notlar'
            }
            
            # Başlıkları ekle
            for col_num, (_, header) in enumerate(headers.items(), 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Örnek satır ekle
            example_row = {
                'name': 'Ahmet Yılmaz',
                'type': 'individual',
                'company_name': '',
                'tax_office': '',
                'tax_number': '',
                'email': 'ahmet.yilmaz@example.com',
                'phone': '0532-123-4567',
                'website': '',
                'notes': 'Örnek müşteri'
            }
            
            for col_num, (field, _) in enumerate(headers.items(), 1):
                cell = ws.cell(row=2, column=col_num, value=example_row[field])
            
            # Kurumsal müşteri örneği
            corp_example = {
                'name': 'Ali Veli',
                'type': 'corporate',
                'company_name': 'ABC Şirketi',
                'tax_office': 'Kadıköy',
                'tax_number': '1234567890',
                'email': 'info@abcsirketi.com',
                'phone': '0216-123-4567',
                'website': 'www.abcsirketi.com',
                'notes': 'Örnek kurumsal müşteri'
            }
            
            for col_num, (field, _) in enumerate(headers.items(), 1):
                cell = ws.cell(row=3, column=col_num, value=corp_example[field])
            
            # Talimatlar sayfası ekle
            ws2 = wb.create_sheet(title="Talimatlar")
            
            ws2.cell(row=1, column=1, value="Müşteri İçe Aktarma Talimatları").font = Font(bold=True, size=14)
            ws2.cell(row=3, column=1, value="Zorunlu Alanlar:").font = Font(bold=True)
            ws2.cell(row=4, column=1, value="- Müşteri Adı (*): Müşteri tam adı")
            ws2.cell(row=5, column=1, value="- Tip (*): 'individual' (bireysel) veya 'corporate' (kurumsal)")
            ws2.cell(row=6, column=1, value="- Email (*): Geçerli email adresi")
            ws2.cell(row=7, column=1, value="- Telefon (*): Telefon numarası")
            
            ws2.cell(row=9, column=1, value="İsteğe Bağlı Alanlar:").font = Font(bold=True)
            ws2.cell(row=10, column=1, value="- Şirket Adı: Kurumsal müşteriler için şirket adı")
            ws2.cell(row=11, column=1, value="- Vergi Dairesi: Kurumsal müşteriler için vergi dairesi")
            ws2.cell(row=12, column=1, value="- Vergi/TC Kimlik No: Vergi numarası veya TC Kimlik numarası")
            ws2.cell(row=13, column=1, value="- Website: Müşteri website adresi")
            ws2.cell(row=14, column=1, value="- Notlar: Müşteri ile ilgili notlar")
            
            ws2.cell(row=16, column=1, value="Önemli Notlar:").font = Font(bold=True)
            ws2.cell(row=17, column=1, value="1. İlk satır başlık satırıdır, değiştirmeyiniz")
            ws2.cell(row=18, column=1, value="2. Zorunlu alanlar (*) işaretli sütunlardır")
            ws2.cell(row=19, column=1, value="3. Aynı email veya telefon numarasına sahip müşteriler güncellenecektir")
            ws2.cell(row=20, column=1, value="4. Tip alanı için 'individual' (bireysel) veya 'corporate' (kurumsal) değerlerini kullanın")
            
            # Talimatlar sayfasında sütun genişliğini ayarla
            ws2.column_dimensions['A'].width = 80
            
            # Sütun genişliklerini ayarla
            for col_num, (_, header) in enumerate(headers.items(), 1):
                column = ws.column_dimensions[chr(64 + col_num)]
                column.width = max(len(header), 20)
            
            # Buffer'a kaydet
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Şablon oluşturma hatası: {str(e)}")
            raise ExcelError(f"Şablon oluşturma başarısız: {str(e)}")
    
    def generate_address_import_template(self) -> bytes:
        """
        Adres içe aktarma için Excel şablonu oluştur.
        
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Adresler"
            
            # Sütun başlıklarını tanımla
            headers = {
                'customer_email': 'Müşteri Email *',
                'title': 'Adres Başlığı *',
                'type': 'Adres Tipi (billing/shipping/other) *',
                'address_line1': 'Adres Satırı 1 *',
                'address_line2': 'Adres Satırı 2',
                'city': 'Şehir *',
                'state': 'İlçe',
                'postal_code': 'Posta Kodu',
                'country': 'Ülke',
                'is_default': 'Varsayılan Adres (TRUE/FALSE)'
            }
            
            # Başlıkları ekle
            for col_num, (_, header) in enumerate(headers.items(), 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Örnek satır ekle
            example_row = {
                'customer_email': 'ahmet.yilmaz@example.com',
                'title': 'Ev Adresi',
                'type': 'shipping',
                'address_line1': 'Atatürk Cad. No:123',
                'address_line2': 'Daire: 5',
                'city': 'İstanbul',
                'state': 'Kadıköy',
                'postal_code': '34000',
                'country': 'Türkiye',
                'is_default': 'TRUE'
            }
            
            for col_num, (field, _) in enumerate(headers.items(), 1):
                cell = ws.cell(row=2, column=col_num, value=example_row[field])
            
            # İkinci bir örnek satır ekle
            example_row2 = {
                'customer_email': 'info@abcsirketi.com',
                'title': 'Merkez Ofis',
                'type': 'billing',
                'address_line1': 'İş Kuleleri No:10',
                'address_line2': 'Kat: 15',
                'city': 'İstanbul',
                'state': 'Levent',
                'postal_code': '34330',
                'country': 'Türkiye',
                'is_default': 'TRUE'
            }
            
            for col_num, (field, _) in enumerate(headers.items(), 1):
                cell = ws.cell(row=3, column=col_num, value=example_row2[field])
            
            # Talimatlar sayfası ekle
            ws2 = wb.create_sheet(title="Talimatlar")
            
            ws2.cell(row=1, column=1, value="Adres İçe Aktarma Talimatları").font = Font(bold=True, size=14)
            ws2.cell(row=3, column=1, value="Zorunlu Alanlar:").font = Font(bold=True)
            ws2.cell(row=4, column=1, value="- Müşteri Email (*): Müşterinin email adresi (sistemde kayıtlı olmalı)")
            ws2.cell(row=5, column=1, value="- Adres Başlığı (*): Adres için kısa başlık")
            ws2.cell(row=6, column=1, value="- Adres Tipi (*): 'billing' (fatura), 'shipping' (teslimat) veya 'other' (diğer)")
            ws2.cell(row=7, column=1, value="- Adres Satırı 1 (*): Ana adres bilgisi")
            ws2.cell(row=8, column=1, value="- Şehir (*): Şehir adı")
            
            ws2.cell(row=10, column=1, value="İsteğe Bağlı Alanlar:").font = Font(bold=True)
            ws2.cell(row=11, column=1, value="- Adres Satırı 2: Ek adres bilgisi")
            ws2.cell(row=12, column=1, value="- İlçe: İlçe veya semt adı")
            ws2.cell(row=13, column=1, value="- Posta Kodu: Posta kodu")
            ws2.cell(row=14, column=1, value="- Ülke: Ülke adı (varsayılan: Türkiye)")
            ws2.cell(row=15, column=1, value="- Varsayılan Adres: 'TRUE' veya 'FALSE' (varsayılan: FALSE)")
            
            ws2.cell(row=17, column=1, value="Önemli Notlar:").font = Font(bold=True)
            ws2.cell(row=18, column=1, value="1. İlk satır başlık satırıdır, değiştirmeyiniz")
            ws2.cell(row=19, column=1, value="2. Zorunlu alanlar (*) işaretli sütunlardır")
            ws2.cell(row=20, column=1, value="3. Müşteri email adresi sisteminizde kayıtlı olmalıdır")
            ws2.cell(row=21, column=1, value="4. Adres tipi için 'billing' (fatura), 'shipping' (teslimat) veya 'other' (diğer) değerlerini kullanın")
            ws2.cell(row=22, column=1, value="5. Aynı müşteri ve başlığa sahip adresler güncellenecektir")
            
            # Talimatlar sayfasında sütun genişliğini ayarla
            ws2.column_dimensions['A'].width = 80
            
            # Sütun genişliklerini ayarla
            for col_num, (_, header) in enumerate(headers.items(), 1):
                column = ws.column_dimensions[chr(64 + col_num)]
                column.width = max(len(header), 20)
            
            # Buffer'a kaydet
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Şablon oluşturma hatası: {str(e)}")
            raise ExcelError(f"Şablon oluşturma başarısız: {str(e)}")
    
    def _create_excel_file(self, queryset, fields, headers, sheet_name='Sheet1') -> bytes:
        """
        Excel dosyası oluştur.
        
        Args:
            queryset: Dışa aktarılacak queryset
            fields: Dışa aktarılacak alanlar
            headers: Sütun başlıkları
            sheet_name: Sayfa adı
            
        Returns:
            bytes: Excel dosyası içeriği
        """
        # Veriyi hazırla
        data = []
        for obj in queryset:
            row = {}
            for field in fields:
                if field in ['type', 'is_active', 'is_default']:
                    # Boolean ve choice alanlar için display değeri kullan
                    if field == 'is_active' or field == 'is_default':
                        row[field] = 'Evet' if getattr(obj, field, False) else 'Hayır'
                    else:
                        display_method = f'get_{field}_display'
                        row[field] = getattr(obj, display_method)() if hasattr(obj, display_method) else getattr(obj, field)
                elif '__' in field:
                    # İlişkili alanlar
                    parts = field.split('__')
                    val = obj
                    for part in parts:
                        if val is not None:
                            val = getattr(val, part) if hasattr(val, part) else None
                    row[field] = val
                else:
                    # Normal alanlar
                    row[field] = getattr(obj, field) if hasattr(obj, field) else None
            data.append(row)
        
        # DataFrame oluştur
        df = pd.DataFrame(data)
        
        # Başlıkları uygula
        df = df.rename(columns=headers)
        
        # Excel dosyasına dönüştür
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Başlıkları ekle
        for col_num, (_, header) in enumerate(headers.items(), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # Verileri ekle
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (field, _) in enumerate(headers.items(), 1):
                field_name = list(fields)[col_idx-1]  # headers ve fields eşleşmeli
                value = row_data.get(field_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Tarih alanları için format
                if 'date' in field_name or 'created_at' in field_name or 'updated_at' in field_name:
                    if value:
                        cell.number_format = 'DD.MM.YYYY HH:MM'
                
                # Hücre hizalaması
                if field_name == 'id':
                    cell.alignment = Alignment(horizontal="center")
                elif field_name in ['is_active', 'is_default']:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Sütun genişliklerini ayarla
        for col_idx, (_, header) in enumerate(headers.items(), 1):
            column = ws.column_dimensions[chr(64 + col_idx)]
            max_length = len(header)
            for row_idx in range(2, len(data) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 50)  # 50'de kapat
            column.width = adjusted_width
        
        # Buffer'a kaydet
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    def get_import_statistics(self, results: Dict[str, Any]) -> str:
        """Import sonuçlarının özet istatistiklerini oluştur"""
        success_rate = (results.get('created', 0) + results.get('updated', 0)) / results.get('total', 1) * 100 if results.get('total', 0) > 0 else 0
        
        return f"""
        Import İstatistikleri:
        - Toplam Satır: {results.get('total', 0)}
        - Oluşturulan: {results.get('created', 0)}
        - Güncellenen: {results.get('updated', 0)}
        - Hatalı: {results.get('error_count', 0)}
        - Başarı Oranı: {success_rate:.1f}%
        """