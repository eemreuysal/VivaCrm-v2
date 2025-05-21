"""
Excel import/export functionality for the VivaCRM v2 application.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.text import slugify
from django.utils import timezone
from django.db import transaction
from django.db.models import Model, QuerySet, Manager
from django.core.exceptions import ValidationError
import logging
import io
from typing import Dict, List, Any, Tuple, Optional, Callable, Type, Union
from .excel_result import ImportResult

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Class to handle exporting data to Excel format.
    """
    
    def __init__(self, queryset: QuerySet, fields: List[str] = None, 
                 headers: Dict[str, str] = None, filename: str = None,
                 sheet_name: str = 'Data', formatting: bool = True):
        """
        Initialize the exporter with the data and configuration.
        
        Args:
            queryset: Django queryset to export
            fields: List of model fields to include (if None, all fields)
            headers: Map of field names to display headers
            filename: Name of the file to be downloaded (without extension)
            sheet_name: Name of the worksheet in Excel
            formatting: Whether to apply formatting to the Excel output
        """
        self.queryset = queryset
        self.model = queryset.model
        self.model_name = self.model._meta.verbose_name_plural.title()
        self.fields = fields
        self.headers = headers or {}
        self.filename = filename or f"{slugify(self.model_name)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
        self.sheet_name = sheet_name
        self.formatting = formatting
    
    def _get_field_names(self) -> List[str]:
        """Get the field names to include in the export."""
        if self.fields:
            return self.fields
        
        # Exclude many-to-many and one-to-many relations by default
        return [field.name for field in self.model._meta.get_fields() 
                if not field.is_relation or not field.many_to_many and not field.one_to_many]
    
    def _get_field_values(self, obj: Model) -> Dict[str, Any]:
        """Get the values for each field from an object."""
        result = {}
        for field_name in self._get_field_names():
            # Handle related objects (e.g., foreign keys)
            if '.' in field_name:
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    if value is None:
                        value = None
                        break
                    value = getattr(value, part, None)
                result[field_name] = value
            else:
                # Try to get attribute directly
                try:
                    value = getattr(obj, field_name)
                    # If it's a callable, call it
                    if callable(value) and not isinstance(value, Manager):
                        value = value()
                except (AttributeError, Manager):
                    try:
                        # Try to get it as a dict key
                        value = obj.get(field_name, '')
                    except (AttributeError, TypeError):
                        value = ''
                        
                result[field_name] = value
        
        return result
    
    def _get_header_label(self, field_name: str) -> str:
        """Get the display label for a field."""
        # If custom header provided, use that
        if field_name in self.headers:
            return self.headers[field_name]
        
        # Otherwise try to get from model field
        try:
            return self.model._meta.get_field(field_name).verbose_name.title()
        except:
            # Fall back to formatted field name
            return field_name.replace('_', ' ').title()
    
    def _format_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                          df: pd.DataFrame) -> None:
        """Apply formatting to the Excel worksheet."""
        if not self.formatting:
            return
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Set column width based on max length in column
            max_length = max(
                df[column_title].astype(str).apply(len).max(),
                len(str(column_title))
            ) + 3  # Add padding
            worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_length, 40)
        
        # Format data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Align numbers to right
                value = cell.value
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    def to_excel(self) -> HttpResponse:
        """
        Export the data to an Excel file and return as HttpResponse.
        """
        # Convert queryset to DataFrame
        data = [self._get_field_values(obj) for obj in self.queryset]
        
        # If no data, create empty DataFrame with just the headers
        if not data:
            fields = self._get_field_names()
            data = [{field: "" for field in fields}]
            
        df = pd.DataFrame(data)
        
        # Rename columns to their display labels
        df.rename(columns={field: self._get_header_label(field) for field in df.columns}, inplace=True)
        
        # Create Excel response
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=self.sheet_name)
            
            # Apply formatting
            if self.formatting:
                self._format_worksheet(writer.sheets[self.sheet_name], df)
        
        # Create HTTP response
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.xlsx"'
        
        return response
    
    def to_csv(self) -> HttpResponse:
        """
        Export the data to a CSV file and return as HttpResponse.
        """
        # Convert queryset to DataFrame
        data = [self._get_field_values(obj) for obj in self.queryset]
        
        # If no data, create empty DataFrame with just the headers
        if not data:
            fields = self._get_field_names()
            data = [{field: "" for field in fields}]
            
        df = pd.DataFrame(data)
        
        # Rename columns to their display labels
        df.rename(columns={field: self._get_header_label(field) for field in df.columns}, inplace=True)
        
        # Create CSV response
        output = io.StringIO()
        df.to_csv(output, index=False)
        
        # Create HTTP response
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.csv"'
        
        return response


class ExcelImporter:
    """
    Class to handle importing data from Excel format.
    """
    
    def __init__(self, model: Type[Model], 
                 field_mapping: Dict[str, str] = None,
                 required_fields: List[str] = None,
                 unique_fields: List[str] = None,
                 validators: Dict[str, Callable] = None,
                 defaults: Dict[str, Any] = None,
                 preprocess_function: Callable = None):
        """
        Initialize the importer with configuration.
        
        Args:
            model: Django model class to import data into
            field_mapping: Map of Excel column names to model field names
            required_fields: List of fields that must be present
            unique_fields: List of fields to check for uniqueness
            validators: Dict of field names to validation functions
            defaults: Dict of field names to default values
            preprocess_function: Optional function to preprocess data before creating/updating
        """
        self.model = model
        self.field_mapping = field_mapping or {}
        self.required_fields = required_fields or []
        self.unique_fields = unique_fields or []
        self.validators = validators or {}
        self.defaults = defaults or {}
        self.preprocess_function = preprocess_function
        
    def _normalize_field_name(self, name: str) -> str:
        """Convert Excel header name to normalized field name."""
        # Try to find in field mapping - exact match first
        if name in self.field_mapping:
            return self.field_mapping[name]
            
        # Try case-insensitive match
        for excel_name, field_name in self.field_mapping.items():
            if excel_name.lower() == name.lower():
                return field_name
        
        # Try direct match with model fields
        for field in self.model._meta.get_fields():
            if field.name.lower() == name.lower():
                return field.name
            if hasattr(field, 'verbose_name') and field.verbose_name.lower() == name.lower():
                return field.name
        
        # Normalize the name (replace spaces with underscores, lowercase)
        return name.lower().replace(' ', '_')
    
    def _validate_required_fields(self, df: pd.DataFrame) -> List[str]:
        """Validate that all required fields are present."""
        missing_fields = []
        normalized_columns = [self._normalize_field_name(col) for col in df.columns]
        
        for field in self.required_fields:
            if field not in normalized_columns:
                # Try to get the Excel column name from field mapping
                for excel_name, model_field in self.field_mapping.items():
                    if model_field == field:
                        field = excel_name
                        break
                        
                missing_fields.append(field)
        
        return missing_fields
    
    def _process_row(self, row: pd.Series) -> Dict[str, Any]:
        """Process a row of data and convert to model field values."""
        data = {}
        
        # Add default values
        for field, value in self.defaults.items():
            data[field] = value
        
        # Process each column
        for column, value in row.items():
            field_name = self._normalize_field_name(column)
            
            # Skip if field is not in model
            if not hasattr(self.model, field_name) and not field_name in self.field_mapping.values():
                continue
            
            # Apply validator if available
            if field_name in self.validators:
                try:
                    value = self.validators[field_name](value)
                except ValidationError as e:
                    raise ValidationError(f"Validation error in field '{column}': {str(e)}")
            
            data[field_name] = value
        
        return data
    
    def _get_lookup_kwargs(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Get kwargs for looking up existing records based on unique fields.
        """
        lookup = {}
        for field in self.unique_fields:
            if field in data:
                lookup[field] = data[field]
        return lookup
    
    @transaction.atomic
    def import_data(self, file_obj: Union[str, io.BytesIO], 
                    sheet_name: str = 0,
                    update_existing: bool = True,
                    import_task: 'ImportTask' = None,
                    user=None) -> ImportResult:
        """
        Import data from an Excel file.
        
        Args:
            file_obj: File path or file-like object containing Excel data
            sheet_name: Name or index of the sheet to import
            update_existing: Whether to update existing records
            import_task: Import task for detailed reporting
            user: User instance for tracking who created/updated records
            
        Returns:
            ImportResult instance containing import statistics
        """
        try:
            # Load data from Excel
            # Check file size and use chunk reading if needed
            if hasattr(file_obj, 'size') and file_obj.size > 10 * 1024 * 1024:  # 10MB
                # For large files, read in chunks
                chunks = pd.read_excel(file_obj, sheet_name=sheet_name, chunksize=1000)
                df = pd.concat(chunks, ignore_index=True)
            else:
                df = pd.read_excel(file_obj, sheet_name=sheet_name)
            
            # Remove unnamed columns and handle NaN values
            df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
            df = df.fillna('')
            
            # Validate required fields
            missing_fields = self._validate_required_fields(df)
            if missing_fields:
                raise ValidationError(f"Missing required fields: {', '.join(missing_fields)}")
            
            # Initialize result tracking
            result = ImportResult()
            if import_task:
                result.initialize_reporter(import_task)
            
            # Find any AUTH_USER_MODEL foreign key fields that might need explicit None values
            auth_user_fields = []
            for field in self.model._meta.fields:
                if hasattr(field, 'remote_field') and field.remote_field is not None and \
                   hasattr(field.remote_field, 'model') and field.remote_field.model.__name__ == 'User':
                    auth_user_fields.append(field.name)
            
            # Process each row
            for index, row in df.iterrows():
                row_number = index + 2  # Excel row number (1-based, plus header)
                row_data = row.to_dict()
                
                # Track field-level success
                fields_succeeded = {}
                fields_failed = {}
                partial_data = {}
                
                try:
                    # Extract data for this row
                    data = self._process_row(row)
                    
                    # Set all AUTH_USER_MODEL foreign key fields to None if not provided
                    for field_name in auth_user_fields:
                        if field_name not in data:
                            data[field_name] = None
                            
                    # Apply preprocessing function if provided
                    if self.preprocess_function:
                        data = self.preprocess_function(data)
                    
                    # Validate each field individually for partial success tracking
                    for field_name, value in data.items():
                        try:
                            # Validate field value
                            field = self.model._meta.get_field(field_name)
                            
                            # Skip validation for ForeignKey fields if value is a model instance
                            from django.db.models import ForeignKey
                            if isinstance(field, ForeignKey) and hasattr(value, 'pk'):
                                # Value is already a model instance from validator
                                fields_succeeded[field_name] = True
                                partial_data[field_name] = value
                            else:
                                field.clean(value, None)
                                fields_succeeded[field_name] = True
                                partial_data[field_name] = value
                        except (ValidationError, Exception) as field_error:
                            fields_failed[field_name] = str(field_error)
                            fields_succeeded[field_name] = False
                    
                    # If some fields succeeded but not all, handle as partial success
                    if fields_failed and fields_succeeded:
                        # Only use successfully validated fields
                        filtered_data = {k: v for k, v in data.items() if fields_succeeded.get(k, False)}
                        
                        # Check if we can still create/update with partial data
                        required_fields_ok = all(fields_succeeded.get(f, False) for f in self.required_fields)
                        
                        if required_fields_ok and filtered_data:
                            lookup_kwargs = self._get_lookup_kwargs(filtered_data)
                            if lookup_kwargs and update_existing:
                                obj, created = self.model.objects.update_or_create(
                                    defaults=filtered_data,
                                    **lookup_kwargs
                                )
                                result.add_partial_success(
                                    row_num=row_number,
                                    fields_succeeded=fields_succeeded,
                                    fields_failed=fields_failed,
                                    data=row_data,
                                    instance_id=obj.id
                                )
                            else:
                                # Create with partial data
                                obj = self.model.objects.create(**filtered_data)
                                result.add_partial_success(
                                    row_num=row_number,
                                    fields_succeeded=fields_succeeded,
                                    fields_failed=fields_failed,
                                    data=row_data,
                                    instance_id=obj.id
                                )
                        else:
                            # Can't proceed with partial data - required fields missing
                            result.add_error(
                                row_num=row_number,
                                error=f"Missing required fields: {', '.join(fields_failed.keys())}",
                                data=row_data
                            )
                    else:
                        # All fields succeeded - normal processing
                        lookup_kwargs = self._get_lookup_kwargs(data)
                        if lookup_kwargs and update_existing:
                            obj, created = self.model.objects.update_or_create(
                                defaults=data,
                                **lookup_kwargs
                            )
                            result.add_success(
                                row_num=row_number,
                                instance_id=obj.id,
                                is_created=created,
                                data=row_data,
                                fields_updated=list(data.keys())
                            )
                        else:
                            # Create new record
                            obj = self.model.objects.create(**data)
                            result.add_success(
                                row_num=row_number,
                                instance_id=obj.id,
                                is_created=True,
                                data=row_data,
                                fields_updated=list(data.keys())
                            )
                        
                except Exception as e:
                    logger.error(f"Error importing row {row_number}: {str(e)}")
                    result.add_error(
                        row_num=row_number,
                        error=str(e),
                        data=row_data
                    )
            
            # Finalize the import
            if result.reporter:
                result.finalize()
                
            return result
            
        except Exception as e:
            logger.error(f"Error importing Excel file: {str(e)}")
            raise ValidationError(f"Error importing Excel file: {str(e)}")


def export_queryset_excel(queryset, fields=None, headers=None, filename=None, 
                         sheet_name=None, formatting=True):
    """
    Utility function to export a queryset to Excel.
    """
    exporter = ExcelExporter(
        queryset=queryset,
        fields=fields,
        headers=headers,
        filename=filename,
        sheet_name=sheet_name or queryset.model._meta.verbose_name_plural.title(),
        formatting=formatting
    )
    return exporter.to_excel()


def export_queryset_csv(queryset, fields=None, headers=None, filename=None):
    """
    Utility function to export a queryset to CSV.
    """
    exporter = ExcelExporter(
        queryset=queryset,
        fields=fields,
        headers=headers,
        filename=filename
    )
    return exporter.to_csv()


def generate_template_excel(model, fields=None, headers=None, filename=None):
    """
    Generate an Excel template for importing data into a model.
    
    Args:
        model: Django model class
        fields: List of field names to include
        headers: Dict mapping field names to display names
        filename: Name of the file to be downloaded (without extension)
        
    Returns:
        HttpResponse with Excel file
    """
    # Determine the fields to include
    if fields is None:
        fields = [field.name for field in model._meta.get_fields() 
                 if not field.is_relation or field.many_to_one]
    
    # Create headers if not provided
    if headers is None:
        headers = {}
        for field_name in fields:
            try:
                field = model._meta.get_field(field_name)
                headers[field_name] = field.verbose_name.title()
            except:
                headers[field_name] = field_name.replace('_', ' ').title()
    
    # Create empty DataFrame with headers
    df = pd.DataFrame(columns=[headers.get(field, field) for field in fields])
    
    # Set filename if not provided
    if filename is None:
        filename = f"{model._meta.verbose_name}_template_{timezone.now().strftime('%Y%m%d')}"
    
    # Create Excel response
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Template")
        
        # Apply formatting to the template
        worksheet = writer.sheets["Template"]
        
        # Format header row
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Set column width based on header length
            worksheet.column_dimensions[get_column_letter(col_num)].width = max(len(column_title) + 5, 15)
    
    # Create HTTP response
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    return response