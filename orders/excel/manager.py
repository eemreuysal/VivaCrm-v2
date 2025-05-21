"""
Order Excel işlemlerini yöneten facade sınıfı.
Facade Pattern: Karmaşık alt sistemlere basit bir arayüz sağlar.
"""
from typing import Optional, Dict, Any, Union
from pathlib import Path
import logging
import pandas as pd
import random
import string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone

from ..services import OrderImportService
from core.excel.chunked import ChunkedExcelReader, MemoryEfficientExcelWriter
from core.excel.exceptions import ExcelError

logger = logging.getLogger(__name__)


class OrderExcelManager:
    """
    Sipariş Excel işlemlerini yöneten ana sınıf.
    Import, export ve diğer Excel işlemlerini tek bir noktadan yönetir.
    """
    
    def import_orders_excel(
        self,
        file_obj,
        update_existing: bool = False,
        user = None
    ) -> Dict[str, Any]:
        """
        Excel dosyasından sipariş verilerini içe aktarır.
        
        Args:
            file_obj: Yüklenen Excel dosyası
            update_existing: Mevcut siparişleri güncelle
            user: İşlemi yapan kullanıcı
            
        Returns:
            dict: Sonuç istatistikleri ve hata mesajları
        """
        try:
            return OrderImportService.import_orders_excel(
                file_obj,
                update_existing=update_existing,
                user=user
            )
                
        except ExcelError as e:
            logger.error(f"Excel import hatası: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Beklenmeyen import hatası: {str(e)}")
            raise ExcelError(f"Import işlemi başarısız: {str(e)}")
    
    def generate_order_template(self) -> bytes:
        """
        Sipariş içe aktarma için Excel şablonu oluşturur.
        
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Siparişler"
            
            # Sütun başlıklarını tanımla
            headers = [
                "SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "EYALET", "ŞEHIR", 
                "SKU", "GTIN", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT", "BIRIM INDIRIM", "SATIR FIYAT"
            ]
            
            # Sütun başlıklarını ekle
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Örnek verileri ekle
            sample_data = [
                [
                    "31.01.2025 23:33", "113-9779299-0493815", "BRİTTANY BULLOCK", "TX", "HOUSTON",
                    "4D-BSW8-HERN", "8684612401053", "Viva Maison Waffle Hand Towel Sets for Bathroom - Soft, Quick Dry, Thin, Lightweight, 100% Turkish Cotton Towels 20x35 (Sage, Set of 4 Hand Towels)",
                    "1", "59,00", "0,00", "59,00"
                ],
                [
                    "31.01.2025 23:23", "113-4240755-5747417", "AMAZON - AMAZON.COM", "LA", "SHREVEPORT",
                    "4R-WTQG-V8NB", "8684612401114", "Viva Maison Waffle Towels, Luxury 2 Pcs Bath & Hand Towel Sets for Bathroom, Soft Quick Dry Thin 100% Turkish Cotton Bath Towels (Caramel, Set of 2 Mix Towels)",
                    "2", "45,00", "0,00", "90,00"
                ]
            ]
            
            # Örnek verileri ekle
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Sayısal sütunlar için sağa hizalama
                    if col_idx in [9, 10, 11, 12]:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            
            # Talimatlar sayfası ekle
            ws2 = wb.create_sheet(title="Talimatlar")
            
            ws2.cell(row=1, column=1, value="Sipariş İçe Aktarma Talimatları").font = Font(bold=True, size=14)
            ws2.cell(row=3, column=1, value="Gerekli Alanlar:").font = Font(bold=True)
            ws2.cell(row=4, column=1, value="- SIPARIŞ TARIHI VE SAATI: GG.AA.YYYY SS:DD formatında (örn: 31.01.2025 23:33)")
            ws2.cell(row=5, column=1, value="- SIPARIŞ NO: Benzersiz sipariş numarası")
            ws2.cell(row=6, column=1, value="- MÜŞTERI ISMI: Tam müşteri adı")
            ws2.cell(row=7, column=1, value="- SKU: Ürün kodu")
            ws2.cell(row=8, column=1, value="- ÜRÜN ISMI: Ürün adı")
            ws2.cell(row=9, column=1, value="- ADET: Sipariş edilen ürün adedi")
            ws2.cell(row=10, column=1, value="- BIRIM FIYAT: Ürün birim fiyatı (virgül ile ayrılmış, örn: 59,00)")
            
            ws2.cell(row=12, column=1, value="İsteğe Bağlı Alanlar:").font = Font(bold=True)
            ws2.cell(row=13, column=1, value="- EYALET: Müşteri eyaleti/ili")
            ws2.cell(row=14, column=1, value="- ŞEHIR: Müşteri şehri/ilçesi")
            ws2.cell(row=15, column=1, value="- GTIN: Ürün barkod numarası")
            ws2.cell(row=16, column=1, value="- BIRIM INDIRIM: Ürüne uygulanan indirim (virgül ile ayrılmış, örn: 10,00)")
            ws2.cell(row=17, column=1, value="- SATIR FIYAT: Toplam satır tutarı ((ADET * BIRIM FIYAT) - BIRIM INDIRIM)")
            
            ws2.cell(row=19, column=1, value="Önemli Notlar:").font = Font(bold=True)
            ws2.cell(row=20, column=1, value="1. İlk satır başlık satırıdır, değiştirmeyiniz")
            ws2.cell(row=21, column=1, value="2. Her satır bir sipariş kalemi olarak işlenir")
            ws2.cell(row=22, column=1, value="3. Aynı sipariş numarasına ait satırlar tek bir sipariş olarak gruplanır")
            ws2.cell(row=23, column=1, value="4. Sayısal değerler için ondalık ayırıcı olarak virgül (,) kullanınız")
            
            # Talimatlar sayfasında sütun genişliğini ayarla
            ws2.column_dimensions['A'].width = 80
            
            # Ana sayfada sütun genişliklerini otomatik ayarla
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = ws.column_dimensions[chr(64 + col)]
                for cell in ws[chr(64 + col)]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 50'de kapat (çok geniş olmasın)
                column.width = adjusted_width
            
            # Buffer'a kaydet
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Şablon oluşturma hatası: {str(e)}")
            raise ExcelError(f"Şablon oluşturma başarısız: {str(e)}")
            
    def generate_order_report(self, queryset, start_date=None, end_date=None, **options) -> bytes:
        """
        Siparişler için Excel raporu oluşturur.
        
        Args:
            queryset: Raporlanacak siparişler
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            **options: Ek seçenekler
            
        Returns:
            bytes: Excel dosya içeriği
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sipariş Raporu"
            
            # Rapor başlığı
            report_title = "Sipariş Raporu"
            if start_date:
                report_title += f" ({start_date.strftime('%d.%m.%Y')}"
                if end_date:
                    report_title += f" - {end_date.strftime('%d.%m.%Y')}"
                report_title += ")"
            
            ws.cell(row=1, column=1, value=report_title).font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f"Oluşturulma Tarihi: {timezone.now().strftime('%d.%m.%Y %H:%M')}")
            
            # Başlıklar
            headers = [
                "Sipariş No", "Müşteri", "Tarih", "Durum", "Ödeme Durumu", 
                "Ara Toplam", "Kargo", "İndirim", "KDV", "Toplam", "Kargo Tarihi",
                "Segment", "Sorumlu"
            ]
            
            # Başlıkları ekle
            header_row = 4
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
            
            # Verileri ekle
            row_idx = header_row + 1
            for order in queryset:
                ws.cell(row=row_idx, column=1, value=order.order_number)
                ws.cell(row=row_idx, column=2, value=order.customer.name if order.customer else "")
                ws.cell(row=row_idx, column=3, value=order.order_date)
                ws.cell(row=row_idx, column=4, value=order.get_status_display())
                ws.cell(row=row_idx, column=5, value=order.get_payment_status_display())
                ws.cell(row=row_idx, column=6, value=float(order.subtotal))
                ws.cell(row=row_idx, column=7, value=float(order.shipping_cost))
                ws.cell(row=row_idx, column=8, value=float(order.discount_amount))
                ws.cell(row=row_idx, column=9, value=float(order.tax_amount))
                ws.cell(row=row_idx, column=10, value=float(order.total_amount))
                ws.cell(row=row_idx, column=11, value=order.shipping_date)
                ws.cell(row=row_idx, column=12, value=order.segment)
                ws.cell(row=row_idx, column=13, value=order.owner.get_full_name() if order.owner else "")
                
                # Formatları ayarla
                # Tarih formatı
                ws.cell(row=row_idx, column=3).number_format = 'DD.MM.YYYY'
                ws.cell(row=row_idx, column=11).number_format = 'DD.MM.YYYY'
                
                # Para formatı
                for i in range(6, 11):
                    ws.cell(row=row_idx, column=i).number_format = '#,##0.00 ₺'
                
                row_idx += 1
            
            # Toplamlar
            total_row = row_idx + 1
            ws.cell(row=total_row, column=1, value="TOPLAM:").font = Font(bold=True)
            
            # Toplamları hesapla ve ekle
            total_subtotal = sum(float(order.subtotal) for order in queryset)
            total_shipping = sum(float(order.shipping_cost) for order in queryset)
            total_discount = sum(float(order.discount_amount) for order in queryset)
            total_tax = sum(float(order.tax_amount) for order in queryset)
            total_amount = sum(float(order.total_amount) for order in queryset)
            
            ws.cell(row=total_row, column=6, value=total_subtotal).font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=total_shipping).font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=total_discount).font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=total_tax).font = Font(bold=True)
            ws.cell(row=total_row, column=10, value=total_amount).font = Font(bold=True)
            
            # Toplam formatları
            for i in range(6, 11):
                ws.cell(row=total_row, column=i).number_format = '#,##0.00 ₺'
            
            # Sütun genişliklerini ayarla
            column_widths = {
                1: 20,  # Sipariş No
                2: 30,  # Müşteri
                3: 15,  # Tarih
                4: 15,  # Durum
                5: 15,  # Ödeme Durumu
                6: 15,  # Ara Toplam
                7: 15,  # Kargo
                8: 15,  # İndirim
                9: 15,  # KDV
                10: 15, # Toplam
                11: 15, # Kargo Tarihi
                12: 15, # Segment
                13: 20, # Sorumlu
            }
            
            for col_num, width in column_widths.items():
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # Buffer'a kaydet
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Rapor oluşturma hatası: {str(e)}")
            raise ExcelError(f"Rapor oluşturma başarısız: {str(e)}")
            
    def get_import_statistics(self, results: Dict[str, Any]) -> str:
        """Import sonuçlarının özet istatistiklerini oluştur"""
        success_rate = (results['created'] + results['updated']) / results['total'] * 100 if results['total'] > 0 else 0
        
        return f"""
        Import İstatistikleri:
        - Toplam Satır: {results['total']}
        - Oluşturulan: {results['created']}
        - Güncellenen: {results['updated']}
        - Hatalı: {results['error_count']}
        - Başarı Oranı: {success_rate:.1f}%
        """