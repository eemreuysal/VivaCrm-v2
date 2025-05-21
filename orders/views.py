from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.views.generic.edit import FormMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.db.models import Q, Count, Sum, Avg
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from datetime import timedelta
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from customers.models import Customer, Address
from products.models import Product
from .models import Order, OrderItem, Payment, Shipment
from .forms import OrderForm, OrderItemForm, PaymentForm, ShipmentForm, OrderSearchForm
from .forms_excel import ExcelImportForm
# Ana excel.py dosyasından direk import
from orders.excel_enhanced import enhanced_import_orders_excel 
from orders.excel import generate_order_template, import_orders_excel
from .excel_validators import OrderExcelValidator
from core.excel_errors import ExcelErrorHandler

logger = logging.getLogger(__name__)


class OrderListView(LoginRequiredMixin, FormMixin, ListView):
    model = Order
    template_name = 'orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 20
    form_class = OrderSearchForm
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Add select_related for performance
        queryset = queryset.select_related('customer', 'billing_address', 'shipping_address', 'owner')
        queryset = queryset.prefetch_related('invoices')
        
        form = self.get_form()
        
        if form.is_valid():
            query = form.cleaned_data.get('query')
            status = form.cleaned_data.get('status')
            payment_status = form.cleaned_data.get('payment_status')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            customer = form.cleaned_data.get('customer')
            min_amount = form.cleaned_data.get('min_amount')
            max_amount = form.cleaned_data.get('max_amount')
            sort_by = form.cleaned_data.get('sort_by', 'order_date')
            sort_dir = form.cleaned_data.get('sort_dir', 'desc')
            
            if query:
                queryset = queryset.filter(
                    Q(order_number__icontains=query) | 
                    Q(customer__name__icontains=query) | 
                    Q(customer__company_name__icontains=query)
                )
            
            if status:
                queryset = queryset.filter(status=status)
                
            if payment_status:
                queryset = queryset.filter(payment_status=payment_status)
                
            if date_from:
                queryset = queryset.filter(order_date__gte=date_from)
                
            if date_to:
                # Add one day to include the end date
                queryset = queryset.filter(order_date__lte=date_to)
                
            if customer:
                queryset = queryset.filter(customer=customer)
                
            if min_amount:
                queryset = queryset.filter(total_amount__gte=min_amount)
                
            if max_amount:
                queryset = queryset.filter(total_amount__lte=max_amount)
            
            # Apply sorting
            if sort_dir == 'desc':
                sort_by = f'-{sort_by}'
            
            queryset = queryset.order_by(sort_by)
        else:
            # Default sorting if no valid form
            queryset = queryset.order_by('-order_date')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        
        # Calculate statistics
        context['total_orders'] = Order.objects.count()
        context['new_orders'] = Order.objects.filter(
            order_date__gte=now - timedelta(days=30)
        ).count()
        
        # Revenue calculations
        total_revenue = Order.objects.filter(
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        context['total_revenue'] = total_revenue
        
        # Last month revenue
        last_month_start = now - timedelta(days=60)
        last_month_end = now - timedelta(days=30)
        last_month_revenue = Order.objects.filter(
            status='completed',
            order_date__gte=last_month_start,
            order_date__lt=last_month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # This month revenue
        this_month_revenue = Order.objects.filter(
            status='completed',
            order_date__gte=now - timedelta(days=30)
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Calculate revenue change
        if last_month_revenue > 0:
            context['revenue_change'] = ((this_month_revenue - last_month_revenue) / last_month_revenue) * 100
        else:
            context['revenue_change'] = 100 if this_month_revenue > 0 else 0
        
        # Pending orders
        context['pending_orders'] = Order.objects.filter(status='pending').count()
        
        # Average order value
        avg_order = Order.objects.filter(
            status='completed'
        ).aggregate(avg_value=Avg('total_amount'))
        context['avg_order_value'] = avg_order['avg_value'] or 0
        
        # Add sorting URLs
        current_params = self.request.GET.copy()
        sort_by = self.request.GET.get('sort_by', 'order_date')
        sort_dir = self.request.GET.get('sort_dir', 'desc')
        
        # Create sort URLs
        context['sort_urls'] = {}
        for field in ['order_number', 'customer', 'order_date', 'total_amount', 'status', 'payment_status']:
            params = current_params.copy()
            params['sort_by'] = field
            
            # Toggle direction if same field
            if sort_by == field:
                params['sort_dir'] = 'asc' if sort_dir == 'desc' else 'desc'
            else:
                params['sort_dir'] = 'desc' if field == 'order_date' else 'asc'
                
            context['sort_urls'][field] = '?' + params.urlencode()
        
        context['sort_by'] = sort_by
        context['sort_dir'] = sort_dir
        
        return context
    
    def get_initial(self):
        return {
            'query': self.request.GET.get('query', ''),
            'status': self.request.GET.get('status', ''),
            'payment_status': self.request.GET.get('payment_status', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'customer': self.request.GET.get('customer', ''),
            'min_amount': self.request.GET.get('min_amount', ''),
            'max_amount': self.request.GET.get('max_amount', ''),
            'sort_by': self.request.GET.get('sort_by', 'order_date'),
            'sort_dir': self.request.GET.get('sort_dir', 'desc'),
        }


class OrderDetailView(LoginRequiredMixin, DetailView):
    model = Order
    template_name = 'orders/order_detail.html'
    context_object_name = 'order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_item_form'] = OrderItemForm(order=self.object)
        context['payment_form'] = PaymentForm(order=self.object)
        context['shipment_form'] = ShipmentForm()
        return context


class OrderCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_form.html'
    success_message = _("Sipariş başarıyla oluşturuldu.")
    
    def form_valid(self, form):
        # Set the owner to the current user if not specified
        if not form.cleaned_data.get('owner'):
            form.instance.owner = self.request.user
        return super().form_valid(form)
    
    def get_success_url(self):
        messages.success(self.request, self.success_message)
        return reverse('orders:order-detail', kwargs={'pk': self.object.pk})


class OrderUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_form.html'
    success_message = _("Sipariş başarıyla güncellendi.")
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.object.pk})


class OrderDeleteView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Order
    template_name = 'orders/order_confirm_delete.html'
    success_url = reverse_lazy('orders:order-list')
    success_message = _("Sipariş başarıyla silindi.")


class AjaxableResponseMixin:
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        # We make sure to call the parent's form_valid() method because
        # it might do some processing (in the case of CreateView, it will
        # call form.save() for example).
        response = super().form_valid(form)
        if self.request.is_ajax():
            data = {
                'success': True,
                'message': self.success_message,
            }
            return JsonResponse(data)
        else:
            return response


# OrderItem Views
class OrderItemCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = OrderItem
    form_class = OrderItemForm
    template_name = 'orders/orderitem_form.html'
    success_message = _("Ürün başarıyla eklendi.")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        self.order = get_object_or_404(Order, pk=self.kwargs['order_pk'])
        kwargs['order'] = self.order
        return kwargs
    
    def form_valid(self, form):
        form.instance.order = self.order
        
        # Use product price and tax if not specified
        if form.instance.product and not form.instance.unit_price:
            if form.instance.product.discount_price and form.instance.product.discount_price > 0:
                form.instance.unit_price = form.instance.product.discount_price
            else:
                form.instance.unit_price = form.instance.product.price
            
        if form.instance.product and not form.instance.tax_rate:
            form.instance.tax_rate = form.instance.product.tax_rate
            
        return super().form_valid(form)
    
    def get_success_url(self):
        messages.success(self.request, self.success_message)
        return reverse('orders:order-detail', kwargs={'pk': self.order.pk})


class OrderItemUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = OrderItem
    form_class = OrderItemForm
    template_name = 'orders/orderitem_form.html'
    success_message = _("Ürün başarıyla güncellendi.")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['order'] = self.object.order
        return kwargs
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.object.order.pk})


class OrderItemDeleteView(LoginRequiredMixin, DeleteView):
    model = OrderItem
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order_pk = self.object.order.pk
        self.object.delete()
        
        if request.is_ajax():
            return JsonResponse({
                'success': True,
                'message': _("Ürün başarıyla silindi.")
            })
        else:
            messages.success(request, _("Ürün başarıyla silindi."))
            return HttpResponseRedirect(reverse('orders:order-detail', kwargs={'pk': order_pk}))
    
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


# Payment Views
class PaymentCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'orders/payment_form.html'
    success_message = _("Ödeme başarıyla eklendi.")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        self.order = get_object_or_404(Order, pk=self.kwargs['order_pk'])
        kwargs['order'] = self.order
        return kwargs
    
    def form_valid(self, form):
        form.instance.order = self.order
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.order.pk})


class PaymentUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'orders/payment_form.html'
    success_message = _("Ödeme başarıyla güncellendi.")
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['order'] = self.object.order
        return kwargs
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.object.order.pk})


class PaymentDeleteView(LoginRequiredMixin, DeleteView):
    model = Payment
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order_pk = self.object.order.pk
        self.object.delete()
        
        if request.is_ajax():
            return JsonResponse({
                'success': True,
                'message': _("Ödeme başarıyla silindi.")
            })
        else:
            messages.success(request, _("Ödeme başarıyla silindi."))
            return HttpResponseRedirect(reverse('orders:order-detail', kwargs={'pk': order_pk}))
    
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


# Shipment Views
class ShipmentCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Shipment
    form_class = ShipmentForm
    template_name = 'orders/shipment_form.html'
    success_message = _("Kargo bilgisi başarıyla eklendi.")
    
    def form_valid(self, form):
        order = get_object_or_404(Order, pk=self.kwargs['order_pk'])
        form.instance.order = order
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.kwargs['order_pk']})


class ShipmentUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Shipment
    form_class = ShipmentForm
    template_name = 'orders/shipment_form.html'
    success_message = _("Kargo bilgisi başarıyla güncellendi.")
    
    def get_success_url(self):
        return reverse('orders:order-detail', kwargs={'pk': self.object.order.pk})


class ShipmentDeleteView(LoginRequiredMixin, DeleteView):
    model = Shipment
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        order_pk = self.object.order.pk
        self.object.delete()
        
        if request.is_ajax():
            return JsonResponse({
                'success': True,
                'message': _("Kargo bilgisi başarıyla silindi.")
            })
        else:
            messages.success(request, _("Kargo bilgisi başarıyla silindi."))
            return HttpResponseRedirect(reverse('orders:order-detail', kwargs={'pk': order_pk}))
    
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


@login_required
def generate_order_template_view(request):
    """
    Generate and download a template for order imports
    """
    return generate_order_template()


@method_decorator(login_required, name='dispatch')
class OrderImportView(FormView):
    """
    View for importing orders from Excel.
    """
    template_name = 'orders/order_import.html'
    form_class = ExcelImportForm
    success_url = reverse_lazy('orders:order-list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
    
    def get(self, request, *args, **kwargs):
        # Eğer API template'i isterse onu göster
        if 'api' in request.GET:
            return render(request, 'orders/order_import_api.html')
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        file_obj = self.request.FILES['excel_file']
        update_existing = form.cleaned_data.get('update_existing', False)
        
        # Debug logging
        logger.info(f"Excel import with file: {file_obj.name}")
        logger.info(f"Update existing: {update_existing}")
        
        # Check if this is an AJAX request
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        try:
            # Import the enhanced Excel function
            from .excel_enhanced import enhanced_import_orders_excel
            
            # Use the enhanced import function with user tracking
            result = enhanced_import_orders_excel(file_obj, user=self.request.user, update_existing=update_existing)
            
            # Handle errors
            if result['error_count'] > 0:
                if is_ajax:
                    return JsonResponse({
                        'success': result['created'] > 0,
                        'errors': result['error_rows'],
                        'error_count': result['error_count'],
                        'created_count': result['created'],
                        'updated_count': result['updated']
                    }, status=200 if result['created'] > 0 else 400)
                
                # Store error data in session for potential export
                self.request.session['error_data'] = result['error_rows']
                
                context = {
                    'errors': result['error_rows'],
                    'error_count': result['error_count'],
                    'created_count': result['created'],
                    'updated_count': result['updated'],
                    'form': form
                }
                return render(self.request, 'orders/order_import_results.html', context)
            
            # Success response
            success_message = f"İçe aktarma tamamlandı: {result['created']} yeni sipariş oluşturuldu"
            if result['updated'] > 0:
                success_message += f", {result['updated']} sipariş güncellendi"
            success_message += "."
            
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': success_message,
                    'created_count': result['created'],
                    'updated_count': result['updated'],
                    'redirect_url': reverse('orders:order-list')
                })
            
            messages.success(self.request, success_message)
            return super().form_valid(form)
            
        except Exception as e:
            logger.error(f"Error importing orders: {str(e)}")
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': f'Import failed: {str(e)}'
                }, status=500)
            
            messages.error(self.request, f"Import failed: {str(e)}")
            form.add_error(None, str(e))
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        is_ajax = self.request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            errors = {field: [str(error) for error in errors_list] for field, errors_list in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return super().form_invalid(form)
    
    def post(self, request, *args, **kwargs):
        # Hata raporunu indir
        if request.GET.get('export_errors') == '1' and 'error_data' in request.session:
            errors_df = pd.DataFrame(request.session['error_data'])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="siparis_hata_raporu.xlsx"'
            errors_df.to_excel(response, index=False)
            return response
            
        return super().post(request, *args, **kwargs)


@login_required
def generate_order_invoice_pdf(request, pk):
    """Generate PDF invoice for an order."""
    order = get_object_or_404(Order, pk=pk)


@login_required
def export_orders(request):
    """
    Export orders to Excel or CSV format.
    """
    # Get format from URL parameters
    export_format = request.GET.get('format', 'xlsx').lower()
    
    # Get filtered queryset
    queryset = Order.objects.all()
    
    # Add select_related for performance
    queryset = queryset.select_related('customer', 'billing_address', 'shipping_address')
    queryset = queryset.prefetch_related('items__product', 'payments', 'shipments')
    
    # Apply filters from request
    form = OrderSearchForm(request.GET)
    if form.is_valid():
        query = form.cleaned_data.get('query')
        status = form.cleaned_data.get('status')
        payment_status = form.cleaned_data.get('payment_status')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        customer = form.cleaned_data.get('customer')
        
        if query:
            queryset = queryset.filter(
                Q(order_number__icontains=query) |
                Q(customer__first_name__icontains=query) |
                Q(customer__last_name__icontains=query) |
                Q(customer__email__icontains=query)
            )
        
        if status:
            queryset = queryset.filter(status=status)
            
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
            
        if date_from:
            queryset = queryset.filter(order_date__date__gte=date_from)
            
        if date_to:
            queryset = queryset.filter(order_date__date__lte=date_to)
            
        if customer:
            queryset = queryset.filter(customer=customer)
    
    # Create Excel file
    if export_format == 'csv':
        return export_orders_csv(queryset)
    else:
        return export_orders_excel(queryset)


@login_required
def export_orders_excel(queryset):
    """Export orders to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Siparişler"
    
    # Headers
    headers = [
        'Sipariş No', 'Sipariş Tarihi', 'Müşteri', 'Müşteri Email', 'Müşteri Telefon',
        'Durum', 'Ödeme Durumu', 'Ödeme Yöntemi', 'Ürün Toplamı', 'Vergi', 
        'Kargo Ücreti', 'İndirim', 'Toplam Tutar', 'Fatura Adresi', 'Teslimat Adresi'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Write data
    for row_num, order in enumerate(queryset, 2):
        values = [
            order.order_number,
            order.order_date.strftime('%Y-%m-%d %H:%M'),
            str(order.customer),
            order.customer.email,
            order.customer.phone or '',
            order.get_status_display(),
            order.get_payment_status_display(),
            order.payment_method,
            float(order.subtotal),
            float(order.tax_amount),
            float(order.shipping_cost),
            float(order.discount_amount),
            float(order.total_amount),
            str(order.billing_address) if order.billing_address else '',
            str(order.shipping_address) if order.shipping_address else ''
        ]
        
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if col_num in [9, 10, 11, 12, 13]:  # Money columns
                cell.number_format = '#,##0.00'
    
    # Add order items sheet
    ws_items = wb.create_sheet(title="Sipariş Detayları")
    
    # Item headers
    item_headers = [
        'Sipariş No', 'Ürün Kodu', 'Ürün Adı', 'Miktar', 
        'Birim Fiyat', 'Vergi Oranı', 'İndirim', 'Toplam'
    ]
    
    for col_num, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Write order items
    row_num = 2
    for order in queryset:
        for item in order.items.all():
            values = [
                order.order_number,
                item.product.code,
                item.product.name,
                item.quantity,
                float(item.unit_price),
                float(item.tax_rate),
                float(item.discount_amount),
                float(item.total_price)
            ]
            
            for col_num, value in enumerate(values, 1):
                cell = ws_items.cell(row=row_num, column=col_num, value=value)
                if col_num in [5, 7, 8]:  # Money columns
                    cell.number_format = '#,##0.00'
            
            row_num += 1
    
    # Adjust column widths
    for ws_obj in [ws, ws_items]:
        for column in ws_obj.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_obj.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=siparisler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def export_orders_csv(queryset):
    """Export orders to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=siparisler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    df_list = []
    
    for order in queryset:
        for item in order.items.all():
            df_list.append({
                'Sipariş No': order.order_number,
                'Sipariş Tarihi': order.order_date.strftime('%Y-%m-%d %H:%M'),
                'Müşteri': str(order.customer),
                'Müşteri Email': order.customer.email,
                'Müşteri Telefon': order.customer.phone or '',
                'Durum': order.get_status_display(),
                'Ödeme Durumu': order.get_payment_status_display(),
                'Ürün Kodu': item.product.code,
                'Ürün Adı': item.product.name,
                'Miktar': item.quantity,
                'Birim Fiyat': float(item.unit_price),
                'Vergi Oranı': float(item.tax_rate),
                'İndirim': float(item.discount_amount),
                'Satır Toplam': float(item.total_price),
                'Kargo Ücreti': float(order.shipping_cost),
                'Sipariş Toplamı': float(order.total_amount)
            })
    
    df = pd.DataFrame(df_list)
    df.to_csv(response, index=False, encoding='utf-8-sig')
    
    return response


@login_required
def generate_order_invoice_pdf_complete(request, pk):
    """Generate PDF invoice for an order."""
    order = get_object_or_404(Order, pk=pk)
    from orders.pdf import generate_invoice_pdf
    return generate_invoice_pdf(order)