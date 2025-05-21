"""
Orders modülü için servis katmanı.
Business logic kod tekrarını önlemek için merkezi bir yer sağlar.
"""
from decimal import Decimal, ROUND_DOWN, InvalidOperation
from django.utils import timezone
from django.core.cache import cache
from django.utils.text import slugify
from django.db.models import Q, Sum, Avg, Count
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.urls import reverse
import logging
import re
import pandas as pd
import random
import string
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import Order, OrderItem, Payment, Shipment
from customers.models import Customer, Address
from products.models import Product

logger = logging.getLogger(__name__)


class OrderNumberService:
    """
    Sipariş numarası oluşturma ve doğrulama işlemleri için servis.
    """
    
    @staticmethod
    def generate_order_number():
        """
        Benzersiz bir sipariş numarası oluştur.
        
        Returns:
            str: Benzersiz sipariş numarası
        """
        # Format: ORD-YYYYMMDD-XXXX 
        prefix = 'ORD'
        date_part = timezone.now().strftime('%Y%m%d')
        
        # Son siparişi bul
        last_order = Order.objects.filter(
            order_number__startswith=f'{prefix}-{date_part}'
        ).order_by('-order_number').first()
        
        if last_order:
            # Son sipariş numarasını al ve arttır
            try:
                last_sequence = int(last_order.order_number.split('-')[-1])
                new_sequence = last_sequence + 1
            except (ValueError, IndexError):
                # Parse hatası olursa rastgele numara oluştur
                new_sequence = random.randint(1, 9999)
        else:
            new_sequence = 1
            
        return f'{prefix}-{date_part}-{new_sequence:04d}'


class OrderCalculationService:
    """
    Sipariş hesaplama işlemleri için servis.
    """
    
    @staticmethod
    def calculate_order_totals(order):
        """
        Sipariş kalemleri üzerinden tüm fiyat alanlarını hesaplar.
        
        Args:
            order (Order): Hesaplanacak sipariş
            
        Returns:
            Order: Hesaplanmış sipariş
        """
        items = order.items.all()
        
        # Ara toplam hesapla
        subtotal = sum(item.line_total for item in items)
        order.subtotal = Decimal(str(subtotal)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
        
        # KDV hesapla
        tax_amount = sum(item.tax_amount for item in items)
        order.tax_amount = Decimal(str(tax_amount)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
        
        # Toplam hesapla
        total = order.subtotal + order.tax_amount + order.shipping_cost - order.discount_amount
        order.total_amount = max(Decimal('0.00'), Decimal(str(total)).quantize(Decimal('0.01'), rounding=ROUND_DOWN))
        
        # Önbelleği temizle
        cache.delete(f'order_{order.id}')
        cache.delete(f'customer_orders_{order.customer_id}')
        
        return order
    
    @staticmethod
    def get_order_statistics(days=30):
        """
        Sipariş istatistiklerini hesaplar.
        
        Args:
            days (int): Gün sayısı
            
        Returns:
            dict: İstatistikler
        """
        now = timezone.now()
        start_date = now - timedelta(days=days)
        
        # İstatistikleri hesapla
        total_orders = Order.objects.count()
        new_orders = Order.objects.filter(order_date__gte=start_date).count()
        
        # Gelir hesaplamaları
        total_revenue = Order.objects.filter(
            status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Önceki dönem geliri
        previous_period_start = now - timedelta(days=days*2)
        previous_period_end = start_date
        
        previous_revenue = Order.objects.filter(
            status='completed',
            order_date__gte=previous_period_start,
            order_date__lt=previous_period_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Mevcut dönem geliri
        current_revenue = Order.objects.filter(
            status='completed',
            order_date__gte=start_date
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Gelir değişimi
        if previous_revenue > 0:
            revenue_change = ((current_revenue - previous_revenue) / previous_revenue) * 100
        else:
            revenue_change = 100 if current_revenue > 0 else 0
        
        # Bekleyen siparişler
        pending_orders = Order.objects.filter(status='pending').count()
        
        # Ortalama sipariş değeri
        avg_order = Order.objects.filter(
            status='completed'
        ).aggregate(avg_value=Avg('total_amount'))
        
        return {
            'total_orders': total_orders,
            'new_orders': new_orders,
            'total_revenue': total_revenue,
            'current_revenue': current_revenue,
            'previous_revenue': previous_revenue,
            'revenue_change': revenue_change,
            'pending_orders': pending_orders,
            'avg_order_value': avg_order['avg_value'] or 0
        }


class OrderStatusService:
    """
    Sipariş durumu işlemleri için servis.
    """
    
    STATUS_TRANSITIONS = {
        'draft': ['pending', 'cancelled'],
        'pending': ['processing', 'cancelled'],
        'processing': ['shipped', 'cancelled'],
        'shipped': ['delivered', 'cancelled', 'returned'],
        'delivered': ['completed', 'returned'],
        'completed': ['returned'],
        'cancelled': [],
        'returned': []
    }
    
    PAYMENT_STATUS_TRANSITIONS = {
        'pending': ['paid', 'partially_paid', 'cancelled'],
        'partially_paid': ['paid', 'cancelled'],
        'paid': ['refunded'],
        'refunded': [],
        'cancelled': []
    }
    
    @staticmethod
    def get_valid_status_transitions(current_status):
        """
        Mevcut durum için geçerli sonraki durumları döndürür.
        
        Args:
            current_status (str): Mevcut sipariş durumu
            
        Returns:
            list: Geçerli sonraki durumlar listesi
        """
        if current_status not in OrderStatusService.STATUS_TRANSITIONS:
            return []
        return OrderStatusService.STATUS_TRANSITIONS[current_status]
    
    @staticmethod
    def get_valid_payment_status_transitions(current_status):
        """
        Mevcut ödeme durumu için geçerli sonraki durumları döndürür.
        
        Args:
            current_status (str): Mevcut ödeme durumu
            
        Returns:
            list: Geçerli sonraki ödeme durumları listesi
        """
        if current_status not in OrderStatusService.PAYMENT_STATUS_TRANSITIONS:
            return []
        return OrderStatusService.PAYMENT_STATUS_TRANSITIONS[current_status]
    
    @staticmethod
    def is_valid_status_transition(current_status, new_status):
        """
        Durum geçişinin geçerli olup olmadığını kontrol eder.
        
        Args:
            current_status (str): Mevcut sipariş durumu
            new_status (str): Yeni sipariş durumu
            
        Returns:
            bool: Geçiş geçerli ise True, değilse False
        """
        if current_status == new_status:
            return True
        valid_transitions = OrderStatusService.get_valid_status_transitions(current_status)
        return new_status in valid_transitions
    
    @staticmethod
    def is_valid_payment_status_transition(current_status, new_status):
        """
        Ödeme durumu geçişinin geçerli olup olmadığını kontrol eder.
        
        Args:
            current_status (str): Mevcut ödeme durumu
            new_status (str): Yeni ödeme durumu
            
        Returns:
            bool: Geçiş geçerli ise True, değilse False
        """
        if current_status == new_status:
            return True
        valid_transitions = OrderStatusService.get_valid_payment_status_transitions(current_status)
        return new_status in valid_transitions
    
    @staticmethod
    def update_order_status(order, new_status, user=None, notes=None):
        """
        Sipariş durumunu günceller ve gerekli işlemleri yapar.
        
        Args:
            order (Order): Güncellenecek sipariş
            new_status (str): Yeni durum
            user (User, optional): İşlemi yapan kullanıcı
            notes (str, optional): Durum değişikliği hakkında notlar
            
        Returns:
            Order: Güncellenmiş sipariş
            
        Raises:
            ValueError: Geçersiz durum geçişi
        """
        if not OrderStatusService.is_valid_status_transition(order.status, new_status):
            raise ValueError(f"Geçersiz durum geçişi: {order.status} -> {new_status}")
        
        old_status = order.status
        order.status = new_status
        
        # İlgili tarih alanlarını güncelle
        if new_status == 'shipped' and not order.shipping_date:
            order.shipping_date = timezone.now()
        elif new_status == 'delivered' and not order.delivery_date:
            order.delivery_date = timezone.now()
        
        # Notları güncelle
        if notes:
            if order.notes:
                order.notes += f"\n\n{timezone.now().strftime('%d.%m.%Y %H:%M')} - Durum değişikliği ({old_status} -> {new_status}): {notes}"
            else:
                order.notes = f"{timezone.now().strftime('%d.%m.%Y %H:%M')} - Durum değişikliği ({old_status} -> {new_status}): {notes}"
        
        order.save()
        
        # Önbelleği temizle
        cache.delete(f'order_{order.id}')
        
        # Durum değişikliğiyle ilgili başka işlemler
        if new_status == 'shipped':
            # Kargo kaydı oluştur
            if not Shipment.objects.filter(order=order, status='shipped').exists():
                Shipment.objects.create(
                    order=order,
                    carrier="Belirtilmemiş",
                    status='shipped',
                    shipping_date=timezone.now(),
                    notes=notes if notes else "Otomatik oluşturuldu"
                )
        
        logger.info(f"Sipariş durumu güncellendi: {order.order_number} - {old_status} -> {new_status} (Kullanıcı: {user.username if user else 'Sistem'})")
        return order


class OrderService:
    """
    Sipariş temel işlemleri için servis.
    """
    
    @staticmethod
    def prepare_order_for_save(order):
        """
        Siparişi kaydetmeden önce hazırlar.
        
        Args:
            order (Order): Hazırlanacak sipariş
            
        Returns:
            Order: Hazırlanmış sipariş
        """
        # Sipariş numarası yoksa oluştur
        if not order.order_number:
            order.order_number = OrderNumberService.generate_order_number()
        
        # Segment bilgisini belirle
        if order.customer and not order.segment:
            # Müşteri ismi "*** ***" ise FBA, değilse FBM
            if order.customer.name and order.customer.name.strip() == "*** ***":
                order.segment = 'FBA'
            else:
                order.segment = 'FBM'
        
        return order
    
    @staticmethod
    def create_order_from_customer(customer, products=None, status='draft', owner=None):
        """
        Müşteri bilgilerine göre sipariş oluşturur.
        
        Args:
            customer (Customer): Sipariş müşterisi
            products (list, optional): Ürün listesi [(product, quantity, unit_price), ...]
            status (str, optional): Sipariş durumu
            owner (User, optional): Sorumlu kullanıcı
            
        Returns:
            Order: Oluşturulan sipariş
        """
        # Yeni sipariş oluştur
        order = Order(
            customer=customer,
            status=status,
            payment_status='pending',
            owner=owner
        )
        
        # Sipariş özelliklerini hazırla
        order = OrderService.prepare_order_for_save(order)
        
        # Fatura adresi ve teslimat adresi atama
        if customer:
            default_address = Address.objects.filter(customer=customer, is_default=True).first()
            if default_address:
                order.billing_address = default_address
                order.shipping_address = default_address
        
        # Siparişi kaydet
        order.save()
        
        # Ürünleri ekle
        if products:
            for product_data in products:
                product, quantity, unit_price = product_data
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=unit_price,
                    tax_rate=product.tax_rate
                )
            
            # Toplamları hesapla
            order = OrderCalculationService.calculate_order_totals(order)
            order.save()
        
        return order
    
    @staticmethod
    def get_filtered_orders(query_params):
        """
        Filtre parametrelerine göre siparişleri getirir.
        
        Args:
            query_params (dict): Filtre parametreleri
            
        Returns:
            QuerySet: Filtrelenmiş siparişler
        """
        queryset = Order.objects.all()
        
        # İlişkili modelleri optimize et
        queryset = queryset.select_related('customer', 'billing_address', 'shipping_address', 'owner')
        queryset = queryset.prefetch_related('items__product', 'payments', 'shipments')
        
        # Filtreleri uygula
        query = query_params.get('query')
        status = query_params.get('status')
        payment_status = query_params.get('payment_status')
        date_from = query_params.get('date_from')
        date_to = query_params.get('date_to')
        customer_id = query_params.get('customer')
        min_amount = query_params.get('min_amount')
        max_amount = query_params.get('max_amount')
        sort_by = query_params.get('sort_by', 'order_date')
        sort_dir = query_params.get('sort_dir', 'desc')
        
        if query:
            queryset = queryset.filter(
                Q(order_number__icontains=query) | 
                Q(customer__name__icontains=query) | 
                Q(customer__company_name__icontains=query)
            )
        
        if status:
            queryset = queryset.filter(status=status)
            
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
            
        if date_from:
            if isinstance(date_from, str):
                try:
                    date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                except ValueError:
                    pass
            queryset = queryset.filter(order_date__date__gte=date_from)
            
        if date_to:
            if isinstance(date_to, str):
                try:
                    date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                except ValueError:
                    pass
            queryset = queryset.filter(order_date__date__lte=date_to)
            
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)
            
        if min_amount:
            queryset = queryset.filter(total_amount__gte=min_amount)
            
        if max_amount:
            queryset = queryset.filter(total_amount__lte=max_amount)
        
        # Sıralama uygula
        if sort_dir == 'desc':
            sort_by = f'-{sort_by}'
        
        queryset = queryset.order_by(sort_by)
        
        return queryset
    
    @staticmethod
    def get_order_details(order_id):
        """
        Sipariş detaylarını getirir.
        
        Args:
            order_id (int): Sipariş ID
            
        Returns:
            dict: Sipariş detayları ve ilişkili veriler
        """
        try:
            order = Order.objects.select_related(
                'customer', 'owner', 'billing_address', 'shipping_address'
            ).prefetch_related(
                'items__product',
                'payments',
                'shipments',
                'invoices'
            ).get(pk=order_id)
            
            return {
                'order': order,
                'items': order.items.select_related('product').all(),
                'payments': order.payments.all(),
                'shipments': order.shipments.all(),
                'total_paid': order.payments.filter(is_successful=True).aggregate(Sum('amount'))['amount__sum'] or 0,
                'address_count': order.customer.addresses.count() if order.customer else 0
            }
        except Order.DoesNotExist:
            logger.error(f"Sipariş bulunamadı: {order_id}")
            return None


class OrderItemService:
    """
    Sipariş kalemi işlemleri için servis.
    """
    
    @staticmethod
    def add_item_to_order(order, product, quantity, unit_price=None, discount_amount=0):
        """
        Siparişe ürün ekler veya mevcut kalemi günceller.
        
        Args:
            order (Order): Sipariş
            product (Product): Eklenecek ürün
            quantity (int): Miktar
            unit_price (Decimal, optional): Birim fiyat, belirtilmezse ürün fiyatı kullanılır
            discount_amount (Decimal, optional): İndirim tutarı
            
        Returns:
            OrderItem: Eklenen veya güncellenen sipariş kalemi
        """
        # Birim fiyat belirtilmemişse ürün fiyatını kullan
        if unit_price is None:
            unit_price = product.current_price
        
        # Mevcut sipariş kalemini bul
        item = OrderItem.objects.filter(order=order, product=product).first()
        
        if item:
            # Mevcut kalemi güncelle
            item.quantity += quantity
            item.unit_price = unit_price
            item.discount_amount = discount_amount
            item.save()
        else:
            # Yeni kalem oluştur
            item = OrderItem.objects.create(
                order=order,
                product=product,
                quantity=quantity,
                unit_price=unit_price,
                tax_rate=product.tax_rate,
                discount_amount=discount_amount
            )
        
        # Sipariş toplamlarını güncelle
        OrderCalculationService.calculate_order_totals(order)
        order.save()
        
        return item
    
    @staticmethod
    def remove_item_from_order(order, item_id):
        """
        Siparişten ürün kaldırır.
        
        Args:
            order (Order): Sipariş
            item_id (int): Kaldırılacak sipariş kalemi ID
            
        Returns:
            bool: Başarılı ise True, değilse False
        """
        try:
            item = OrderItem.objects.get(id=item_id, order=order)
            item.delete()
            
            # Sipariş toplamlarını güncelle
            OrderCalculationService.calculate_order_totals(order)
            order.save()
            
            return True
        except OrderItem.DoesNotExist:
            return False


class OrderExportService:
    """
    Sipariş dışa aktarma işlemleri için servis.
    """
    
    @staticmethod
    def export_orders_to_excel(queryset):
        """
        Siparişleri Excel formatında dışa aktarır.
        
        Args:
            queryset: Sipariş QuerySet
            
        Returns:
            HttpResponse: Excel dosyası içeren yanıt
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Siparişler"
        
        # Başlıklar
        headers = [
            'Sipariş No', 'Sipariş Tarihi', 'Müşteri', 'Müşteri Email', 'Müşteri Telefon',
            'Durum', 'Ödeme Durumu', 'Ödeme Yöntemi', 'Ürün Toplamı', 'Vergi', 
            'Kargo Ücreti', 'İndirim', 'Toplam Tutar', 'Fatura Adresi', 'Teslimat Adresi'
        ]
        
        # Başlıkları yaz
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
        # Verileri yaz
        for row_num, order in enumerate(queryset, 2):
            values = [
                order.order_number,
                order.order_date.strftime('%Y-%m-%d %H:%M'),
                str(order.customer),
                order.customer.email if order.customer else '',
                order.customer.phone or '',
                order.get_status_display(),
                order.get_payment_status_display(),
                order.get_payment_method_display() if order.payment_method else '',
                float(order.subtotal),
                float(order.tax_amount),
                float(order.shipping_cost),
                float(order.discount_amount),
                float(order.total_amount),
                str(order.billing_address) if order.billing_address else '',
                str(order.shipping_address) if order.shipping_address else ''
            ]
            
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num in [9, 10, 11, 12, 13]:  # Para sütunları
                    cell.number_format = '#,##0.00'
        
        # Sipariş kalemleri sayfası ekle
        ws_items = wb.create_sheet(title="Sipariş Detayları")
        
        # Kalem başlıkları
        item_headers = [
            'Sipariş No', 'Ürün Kodu', 'Ürün Adı', 'Miktar', 
            'Birim Fiyat', 'Vergi Oranı', 'İndirim', 'Toplam'
        ]
        
        for col_num, header in enumerate(item_headers, 1):
            cell = ws_items.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Sipariş kalemlerini yaz
        row_num = 2
        for order in queryset:
            for item in order.items.all():
                values = [
                    order.order_number,
                    item.product.code if hasattr(item.product, 'code') else item.product.sku,
                    item.product.name,
                    item.quantity,
                    float(item.unit_price),
                    float(item.tax_rate),
                    float(item.discount_amount),
                    float(item.line_total + item.tax_amount)
                ]
                
                for col_num, value in enumerate(values, 1):
                    cell = ws_items.cell(row=row_num, column=col_num, value=value)
                    if col_num in [5, 7, 8]:  # Para sütunları
                        cell.number_format = '#,##0.00'
                
                row_num += 1
        
        # Sütun genişliklerini ayarla
        for ws_obj in [ws, ws_items]:
            for column in ws_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Yanıt oluştur
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=siparisler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_orders_to_csv(queryset):
        """
        Siparişleri CSV formatında dışa aktarır.
        
        Args:
            queryset: Sipariş QuerySet
            
        Returns:
            HttpResponse: CSV dosyası içeren yanıt
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=siparisler_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        df_list = []
        
        for order in queryset:
            for item in order.items.all():
                df_list.append({
                    'Sipariş No': order.order_number,
                    'Sipariş Tarihi': order.order_date.strftime('%Y-%m-%d %H:%M'),
                    'Müşteri': str(order.customer),
                    'Müşteri Email': order.customer.email if order.customer else '',
                    'Müşteri Telefon': order.customer.phone or '',
                    'Durum': order.get_status_display(),
                    'Ödeme Durumu': order.get_payment_status_display(),
                    'Ürün Kodu': item.product.code if hasattr(item.product, 'code') else item.product.sku,
                    'Ürün Adı': item.product.name,
                    'Miktar': item.quantity,
                    'Birim Fiyat': float(item.unit_price),
                    'Vergi Oranı': float(item.tax_rate),
                    'İndirim': float(item.discount_amount),
                    'Satır Toplam': float(item.line_total + item.tax_amount),
                    'Kargo Ücreti': float(order.shipping_cost),
                    'Sipariş Toplamı': float(order.total_amount)
                })
        
        df = pd.DataFrame(df_list)
        df.to_csv(response, index=False, encoding='utf-8-sig')
        
        return response


class OrderImportService:
    """
    Excel'den sipariş içe aktarma işlemleri için servis.
    """
    
    @staticmethod
    def validate_excel_file(df):
        """
        Excel dosyasının içeriğini doğrular.
        
        Args:
            df (DataFrame): Doğrulanacak DataFrame
            
        Returns:
            tuple: (is_valid, errors)
        """
        errors = []
        
        # Gerekli sütunları kontrol et
        required_columns = ["SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "SKU", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            errors.append(f"Gerekli sütunlar eksik: {', '.join(missing_columns)}")
            return False, errors
        
        # Veri doğrulaması
        for idx, row in df.iterrows():
            row_num = idx + 2  # +2 çünkü Excel 1'den başlar ve bir başlık satırı var
            
            # Boş hücreleri kontrol et
            for field in ["SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "SKU", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT"]:
                if pd.isna(row[field]) or str(row[field]).strip() == "":
                    errors.append(f"Satır {row_num}: {field} alanı boş olamaz")
            
            # Sayısal değerleri kontrol et
            try:
                if not pd.isna(row["ADET"]):
                    quantity = int(row["ADET"])
                    if quantity <= 0:
                        errors.append(f"Satır {row_num}: ADET 0'dan büyük olmalıdır")
            except:
                errors.append(f"Satır {row_num}: ADET geçerli bir sayı değil: {row['ADET']}")
            
            # Fiyatları kontrol et
            try:
                if not pd.isna(row["BIRIM FIYAT"]):
                    unit_price_str = str(row["BIRIM FIYAT"]).replace(',', '.')
                    unit_price = Decimal(unit_price_str)
                    if unit_price < 0:
                        errors.append(f"Satır {row_num}: BIRIM FIYAT negatif olamaz")
            except:
                errors.append(f"Satır {row_num}: BIRIM FIYAT geçerli bir sayı değil: {row['BIRIM FIYAT']}")
            
            # İndirimi kontrol et
            if "BIRIM INDIRIM" in row and not pd.isna(row["BIRIM INDIRIM"]):
                try:
                    discount_str = str(row["BIRIM INDIRIM"]).replace(',', '.')
                    discount = Decimal(discount_str)
                    if discount < 0:
                        errors.append(f"Satır {row_num}: BIRIM INDIRIM negatif olamaz")
                except:
                    errors.append(f"Satır {row_num}: BIRIM INDIRIM geçerli bir sayı değil: {row['BIRIM INDIRIM']}")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def import_orders_excel(file_obj, update_existing=False, user=None):
        """
        Excel dosyasından sipariş verilerini içe aktarır.
        
        Args:
            file_obj: Yüklenen Excel dosyası
            update_existing (bool): Mevcut siparişleri güncelle
            user (User, optional): İşlemi yapan kullanıcı
            
        Returns:
            dict: Sonuç istatistikleri ve hata mesajları
        """
        try:
            # Excel dosyasını oku
            df = pd.read_excel(file_obj)
            
            # Doğrulama yap
            is_valid, errors = OrderImportService.validate_excel_file(df)
            if not is_valid:
                return {
                    'total': 0,
                    'created': 0,
                    'updated': 0,
                    'error_count': len(errors),
                    'error_rows': [{'row': 'Header', 'error': err} for err in errors]
                }
            
            # Sonuçları başlat
            result = {
                'total': len(df),
                'created': 0,
                'updated': 0,
                'error_count': 0,
                'error_rows': []
            }
            
            # Grupla ve işle
            order_groups = df.groupby("SIPARIŞ NO")
            
            # Her sipariş grubunu işle
            for order_number, order_group in order_groups:
                try:
                    # İlk satırı al
                    first_row = order_group.iloc[0]
                    row_num = first_row.name + 2
                    
                    # Mevcut siparişi kontrol et
                    existing_order = None
                    if update_existing:
                        try:
                            existing_order = Order.objects.get(order_number=order_number)
                        except Order.DoesNotExist:
                            pass
                    
                    # Müşteri al veya oluştur
                    customer_name = first_row["MÜŞTERI ISMI"]
                    customer = OrderImportService._get_or_create_customer(
                        customer_name, 
                        first_row.get('EYALET', ''), 
                        first_row.get('ŞEHIR', '')
                    )
                    
                    # Sipariş tarihini ayrıştır
                    order_date = OrderImportService._parse_order_date(first_row["SIPARIŞ TARIHI VE SAATI"])
                    
                    # Sipariş oluştur veya güncelle
                    if existing_order:
                        order = existing_order
                        order.customer = customer
                        order.order_date = order_date
                    else:
                        # Sipariş notlarını hazırla
                        location_info = []
                        if not pd.isna(first_row.get('EYALET', '')):
                            location_info.append(first_row.get('EYALET', ''))
                        if not pd.isna(first_row.get('ŞEHIR', '')):
                            location_info.append(first_row.get('ŞEHIR', ''))
                        
                        order_notes = "Excel'den içe aktarıldı"
                        if location_info:
                            order_notes += f": {', '.join(location_info)}"
                            
                        order = Order(
                            order_number=order_number,
                            customer=customer,
                            status='pending',  # Varsayılan durumu beklemede
                            payment_status='pending',  # Varsayılan ödeme durumu beklemede
                            order_date=order_date,
                            notes=order_notes,
                            owner=user
                        )
                    
                    # Siparişi kaydet
                    order.save()
                    
                    # Güncelleme ise mevcut kalemleri temizle
                    if existing_order and update_existing:
                        order.items.all().delete()
                    
                    # Sipariş kalemlerini işle
                    success = OrderImportService._process_order_items(order, order_group, result)
                    
                    if success:
                        # Sipariş toplamlarını hesapla
                        OrderCalculationService.calculate_order_totals(order)
                        order.save()
                        
                        if existing_order:
                            result['updated'] += 1
                        else:
                            result['created'] += 1
                    
                except Exception as e:
                    result['error_count'] += 1
                    result['error_rows'].append({
                        'row': row_num,
                        'error': f"Sipariş hatası: {str(e)}"
                    })
            
            # Hata sayısını güncelle
            result['error_count'] = len(result['error_rows'])
            
            return result
            
        except Exception as e:
            # Dosya okuma/format hatalarını yönet
            raise Exception(f"Excel dosyası işlenirken hata oluştu: {str(e)}")
    
    @staticmethod
    def _get_or_create_customer(customer_name, state=None, city=None):
        """
        Müşteri adına göre müşteriyi bulur veya oluşturur.
        
        Args:
            customer_name (str): Müşteri adı
            state (str, optional): Eyalet/İl
            city (str, optional): Şehir/İlçe
            
        Returns:
            Customer: Bulunan veya oluşturulan müşteri
        """
        if pd.isna(customer_name) or not customer_name:
            raise ValueError("Müşteri ismi zorunludur")
        
        # Önce mevcut müşteriyi bul
        customers = Customer.objects.filter(name__iexact=customer_name)
        if customers.exists():
            return customers.first()
        
        # E-posta oluştur
        email_name = re.sub(r'[^\w\s]', '', customer_name.lower())
        email_name = re.sub(r'\s+', '.', email_name.strip())
        
        # E-posta adı boş olmamalı
        if not email_name:
            email_name = "customer"
            
        # Yeni müşteri oluştur
        customer = Customer.objects.create(
            name=customer_name,
            email=f"{email_name}@example.com",  # Geçici e-posta
            phone="",  # Boş telefon
            company_name="",  # Boş şirket adı
            notes=""
        )
        
        # Konum bilgisi notlara ekle
        location_info = []
        if state and not pd.isna(state):
            location_info.append(f"Eyalet: {state}")
        if city and not pd.isna(city):
            location_info.append(f"Şehir: {city}")
        
        if location_info:
            customer.notes = ", ".join(location_info)
            customer.save()
        
        # Adres oluştur
        if (state and not pd.isna(state)) or (city and not pd.isna(city)):
            Address.objects.create(
                customer=customer,
                title="Excel Import Adresi",
                type="shipping",
                city=city if city and not pd.isna(city) else "",
                state=state if state and not pd.isna(state) else "",
                address_line1="Excel import'tan oluşturuldu",
                country="Türkiye",
                is_default=True
            )
        
        return customer
    
    @staticmethod
    def _parse_order_date(order_date_str):
        """
        Sipariş tarihini ayrıştırır.
        
        Args:
            order_date_str (str): Sipariş tarihi string
            
        Returns:
            datetime: Ayrıştırılmış tarih
        """
        try:
            # Önce DD.MM.YYYY HH:MM formatını dene
            order_date = pd.to_datetime(order_date_str, format="%d.%m.%Y %H:%M")
        except:
            # Başarısız olursa pandas varsayılan ayrıştırıcısını kullan
            try:
                order_date = pd.to_datetime(order_date_str)
            except:
                # Son çare olarak şu anki zamanı kullan
                order_date = timezone.now()
                logger.warning(f"Geçersiz tarih formatı: {order_date_str}, varsayılan tarih kullanıldı")
        
        return order_date
    
    @staticmethod
    def _process_order_items(order, order_group, result):
        """
        Sipariş kalemlerini işler.
        
        Args:
            order (Order): Sipariş
            order_group (DataFrame): Sipariş kalemleri DataFrame
            result (dict): Sonuç istatistikleri
            
        Returns:
            bool: Başarılı ise True, değilse False
        """
        success = True
        
        for idx, item_row in order_group.iterrows():
            try:
                item_row_num = idx + 2  # Hata raporlama için
                
                # SKU ve ürün adını al
                sku = item_row["SKU"]
                product_name = item_row["ÜRÜN ISMI"]
                
                if pd.isna(sku) or pd.isna(product_name):
                    raise ValueError("SKU ve ürün ismi zorunludur")
                
                # Miktar ve fiyat değerlerini ayrıştır
                try:
                    quantity = int(item_row["ADET"])
                except:
                    raise ValueError(f"Geçersiz adet: {item_row['ADET']}")
                
                # Fiyatı ayrıştır (Avrupa'dan ABD formatına dönüştür)
                unit_price_str = str(item_row["BIRIM FIYAT"]).replace(',', '.')
                try:
                    try:
                        unit_price = Decimal(unit_price_str)
                    except InvalidOperation:
                        raise ValueError(f"Geçersiz birim fiyat: {item_row['BIRIM FIYAT']}")
                except:
                    raise ValueError(f"Geçersiz birim fiyat: {item_row['BIRIM FIYAT']}")
                
                # İndirimi kontrol et
                discount_amount = Decimal('0.00')
                if "BIRIM INDIRIM" in item_row and not pd.isna(item_row["BIRIM INDIRIM"]):
                    discount_str = str(item_row["BIRIM INDIRIM"]).replace(',', '.')
                    try:
                        try:
                            discount_amount = Decimal(discount_str)
                        except InvalidOperation:
                            result['error_rows'].append({
                                'row': item_row_num,
                                'error': f"Geçersiz indirim tutarı: {item_row['BIRIM INDIRIM']}, 0 olarak alındı"
                            })
                    except:
                        result['error_rows'].append({
                            'row': item_row_num,
                            'error': f"Geçersiz indirim tutarı: {item_row['BIRIM INDIRIM']}, 0 olarak alındı"
                        })
                
                # Ürün bul veya oluştur
                product = OrderImportService._get_or_create_product(
                    sku, 
                    product_name, 
                    unit_price, 
                    item_row.get("GTIN", "")
                )
                
                # Sipariş kalemi oluştur
                unit_price_decimal = Decimal(str(unit_price)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                discount_amount_decimal = Decimal(str(discount_amount)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=unit_price_decimal,
                    discount_amount=discount_amount_decimal,
                    tax_rate=product.tax_rate
                )
                
            except Exception as e:
                success = False
                detailed_error = str(e)
                logger.error(f"Satır {item_row_num} hatası: {detailed_error}")
                result['error_rows'].append({
                    'row': item_row_num,
                    'error': f"Sipariş kalemi hatası: {detailed_error}"
                })
        
        return success
    
    @staticmethod
    def _get_or_create_product(sku, product_name, price, gtin=None):
        """
        SKU'ya göre ürünü bulur veya oluşturur.
        
        Args:
            sku (str): Ürün SKU
            product_name (str): Ürün adı
            price (Decimal): Ürün fiyatı
            gtin (str, optional): GTIN/Barkod
            
        Returns:
            Product: Bulunan veya oluşturulan ürün
        """
        # Önce SKU ile bul
        products = Product.objects.filter(sku=sku)
        if products.exists():
            return products.first()
        
        # GTIN/Barkod ile bul
        if gtin and not pd.isna(gtin):
            gtin_raw = gtin
            if isinstance(gtin_raw, float):
                gtin_value = str(int(gtin_raw))
            else:
                gtin_value = str(gtin_raw)
            
            products_by_barcode = Product.objects.filter(barcode=gtin_value)
            if products_by_barcode.exists():
                return products_by_barcode.first()
        
        # GTIN değerini ayarla
        gtin_value = ""
        if gtin and not pd.isna(gtin):
            gtin_raw = gtin
            if isinstance(gtin_raw, float):
                gtin_value = str(int(gtin_raw))
            else:
                gtin_value = str(gtin_raw)
        
        # Yeni ürün oluştur
        product = Product.objects.create(
            name=product_name,
            sku=sku,
            barcode=gtin_value,
            description="",
            price=price,
            discount_price=0,
            tax_rate=18,  # Varsayılan KDV oranı
            stock=100,  # Varsayılan stok
            is_active=True,
            code=sku  # Code alanını ayarla (zorunlu)
        )
        
        return product


class OrderExcelManager:
    """
    Excel işlemleri için facade pattern implementasyonu.
    Tüm Excel import/export işlemleri için tek bir erişim noktası sağlar.
    """
    
    def __init__(self):
        """
        Excel manager'ını başlat.
        """
        pass
    
    def generate_order_template(self):
        """
        Sipariş import şablonu oluşturur.
        
        Returns:
            HttpResponse: Excel şablonu
        """
        from .excel import generate_order_template
        return generate_order_template()
    
    def import_orders_excel(self, file_obj, user=None, update_existing=False):
        """
        Excel dosyasından siparişleri içe aktarır.
        
        Args:
            file_obj: Excel dosyası
            user: İşlemi yapan kullanıcı
            update_existing: Mevcut siparişleri güncelle
            
        Returns:
            dict: Sonuç istatistikleri
        """
        try:
            # İleri seviye import fonksiyonunu kullan
            from .excel_enhanced import enhanced_import_orders_excel
            return enhanced_import_orders_excel(file_obj, user=user, update_existing=update_existing)
        except ImportError:
            # Fallback olarak standart import fonksiyonunu kullan
            return OrderImportService.import_orders_excel(file_obj, update_existing=update_existing, user=user)
    
    def export_orders_excel(self, queryset, format_type='xlsx'):
        """
        Siparişleri dışa aktarır.
        
        Args:
            queryset: Sipariş QuerySet
            format_type: Dosya formatı ('xlsx' veya 'csv')
            
        Returns:
            HttpResponse: Excel veya CSV dosyası
        """
        if format_type == 'csv':
            return OrderExportService.export_orders_to_csv(queryset)
        else:
            return OrderExportService.export_orders_to_excel(queryset)
    
    def validate_excel_file(self, file_obj):
        """
        Excel dosyasını doğrular.
        
        Args:
            file_obj: Excel dosyası
            
        Returns:
            tuple: (geçerli mi, hatalar)
        """
        try:
            df = pd.read_excel(file_obj)
            return OrderImportService.validate_excel_file(df)
        except Exception as e:
            return False, [f"Dosya okunamadı: {str(e)}"]