"""
Excel import/export functionality for the Orders app.
"""
import pandas as pd
import uuid
import random
import string
import re
import logging
from decimal import Decimal, ROUND_DOWN, InvalidOperation
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from django.http import HttpResponse

# Set up logger
logger = logging.getLogger(__name__)

from .models import Order, OrderItem
from customers.models import Customer, Address
from products.models import Product


def generate_order_number():
    """Generate a unique order number."""
    # Format: ORD-YYYYMMDD-XXXX where XXXX is a random string
    today = timezone.now().strftime('%Y%m%d')
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"ORD-{today}-{random_str}"


def generate_order_template():
    """
    Generate an Excel template for importing orders.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Siparişler"
    
    # Define column headers based on the requested format
    headers = [
        "SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "EYALET", "ŞEHIR", 
        "SKU", "GTIN", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT", "BIRIM INDIRIM", "SATIR FIYAT"
    ]
    
    # Add column headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # Add sample data rows
    sample_data = [
        [
            "31.01.2025 23:33", "113-9779299-0493815", "BRİTTANY BULLOCK", "TX", "HOUSTON",
            "4D-BSW8-HERN", "8684612401053", "Viva Maison Waffle Hand Towel Sets for Bathroom - Soft, Quick Dry, Thin, Lightweight, 100% Turkish Cotton Towels 20x35 (Sage, Set of 4 Hand Towels)",
            "1", "59,00", "0,00", "59,00"
        ],
        [
            "31.01.2025 23:23", "113-4240755-5747417", "AMAZON - AMAZON.COM", "LA", "SHREVEPORT",
            "4R-WTQG-V8NB", "8684612401114", "Viva Maison Waffle Towels, Luxury 2 Pcs Bath & Hand Towel Sets for Bathroom, Soft Quick Dry Thin 100% Turkish Cotton Bath Towels (Caramel, Set of 2 Mix Towels)",
            "2", "45,00", "0,00", "90,00"
        ]
    ]
    
    # Add sample data to the sheet
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # For better readability of the data
            if col_idx in [9, 10, 11, 12]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Add instructions in a new sheet
    ws2 = wb.create_sheet(title="Talimatlar")
    
    ws2.cell(row=1, column=1, value="Sipariş İçe Aktarma Talimatları").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Gerekli Alanlar:").font = Font(bold=True)
    ws2.cell(row=4, column=1, value="- SIPARIŞ TARIHI VE SAATI: GG.AA.YYYY SS:DD formatında (örn: 31.01.2025 23:33)")
    ws2.cell(row=5, column=1, value="- SIPARIŞ NO: Benzersiz sipariş numarası")
    ws2.cell(row=6, column=1, value="- MÜŞTERI ISMI: Tam müşteri adı")
    ws2.cell(row=7, column=1, value="- SKU: Ürün kodu")
    ws2.cell(row=8, column=1, value="- ÜRÜN ISMI: Ürün adı")
    ws2.cell(row=9, column=1, value="- ADET: Sipariş edilen ürün adedi")
    ws2.cell(row=10, column=1, value="- BIRIM FIYAT: Ürün birim fiyatı (virgül ile ayrılmış, örn: 59,00)")
    
    ws2.cell(row=12, column=1, value="İsteğe Bağlı Alanlar:").font = Font(bold=True)
    ws2.cell(row=13, column=1, value="- EYALET: Müşteri eyaleti/ili")
    ws2.cell(row=14, column=1, value="- ŞEHIR: Müşteri şehri/ilçesi")
    ws2.cell(row=15, column=1, value="- GTIN: Ürün barkod numarası")
    ws2.cell(row=16, column=1, value="- BIRIM INDIRIM: Ürüne uygulanan indirim (virgül ile ayrılmış, örn: 10,00)")
    ws2.cell(row=17, column=1, value="- SATIR FIYAT: Toplam satır tutarı ((ADET * BIRIM FIYAT) - BIRIM INDIRIM)")
    
    ws2.cell(row=19, column=1, value="Önemli Notlar:").font = Font(bold=True)
    ws2.cell(row=20, column=1, value="1. İlk satır başlık satırıdır, değiştirmeyiniz")
    ws2.cell(row=21, column=1, value="2. Her satır bir sipariş kalemi olarak işlenir")
    ws2.cell(row=22, column=1, value="3. Aynı sipariş numarasına ait satırlar tek bir sipariş olarak gruplanır")
    ws2.cell(row=23, column=1, value="4. Sayısal değerler için ondalık ayırıcı olarak virgül (,) kullanınız")
    
    # Adjust column widths in the instructions sheet
    ws2.column_dimensions['A'].width = 80
    
    # Auto-adjust column widths in the main sheet
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = ws.column_dimensions[chr(64 + col)]
        for cell in ws[chr(64 + col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 to avoid excessive widths
        column.width = adjusted_width
    
    # Save to response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=siparis_excel_sablonu.xlsx'
    
    return response


def import_orders_excel(file_obj, update_existing=False):
    """
    Import orders from an Excel file with the new format.
    
    Args:
        file_obj: The uploaded Excel file
        update_existing: Whether to update existing orders by order_number
        
    Returns:
        dict: Result statistics and error messages
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_obj)
        
        # Verify the expected columns
        expected_columns = [
            "SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "EYALET", "ŞEHIR", 
            "SKU", "GTIN", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT", "BIRIM INDIRIM", "SATIR FIYAT"
        ]
        
        # Check if all required columns exist
        required_columns = ["SIPARIŞ TARIHI VE SAATI", "SIPARIŞ NO", "MÜŞTERI ISMI", "SKU", "ÜRÜN ISMI", "ADET", "BIRIM FIYAT"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Gerekli sütunlar eksik: {', '.join(missing_columns)}")
        
        # Initialize results
        result = {
            'total': len(df),
            'created': 0,
            'updated': 0,
            'error_count': 0,
            'error_rows': []
        }
        
        # Group rows by order number
        order_groups = df.groupby("SIPARIŞ NO")
        
        # Process each order group
        for order_number, order_group in order_groups:
            try:
                # Process only the first row for order details
                first_row = order_group.iloc[0]
                row_num = first_row.name + 2  # +2 because Excel is 1-indexed and we have a header row
                
                # Check if order already exists
                existing_order = None
                if update_existing:
                    try:
                        existing_order = Order.objects.get(order_number=order_number)
                    except Order.DoesNotExist:
                        pass
                
                # Get or create customer based on name
                customer_name = first_row["MÜŞTERI ISMI"]
                if pd.isna(customer_name) or not customer_name:
                    raise ValueError("Müşteri ismi zorunludur")
                
                # Try to find an existing customer with this name
                customer = None
                customers = Customer.objects.filter(name__iexact=customer_name)
                if customers.exists():
                    customer = customers.first()
                else:
                    # Create a valid email from customer name
                    # Remove special characters and replace spaces with dots
                    email_name = re.sub(r'[^\w\s]', '', customer_name.lower())
                    email_name = re.sub(r'\s+', '.', email_name.strip())
                    
                    # Ensure email name is not empty
                    if not email_name:
                        email_name = "customer"
                        
                    # Create a new customer if not found
                    customer = Customer.objects.create(
                        name=customer_name,
                        email=f"{email_name}@example.com",  # Placeholder email
                        phone="",  # Empty phone
                        company_name="",  # Empty company name
                        notes=f"Eyalet: {first_row.get('EYALET', '')}, Şehir: {first_row.get('ŞEHIR', '')}" if not pd.isna(first_row.get('ŞEHIR', '')) else ""
                    )
                    
                    # Create an address for the customer if we have city/state data
                    if not pd.isna(first_row.get('ŞEHIR', '')) or not pd.isna(first_row.get('EYALET', '')):
                        Address.objects.create(
                            customer=customer,
                            title="Excel Import Adresi",
                            type="shipping",
                            city=first_row.get('ŞEHIR', '') if not pd.isna(first_row.get('ŞEHIR', '')) else "",
                            state=first_row.get('EYALET', '') if not pd.isna(first_row.get('EYALET', '')) else "",
                            address_line1="Excel import'tan oluşturuldu",
                            country="Türkiye",
                            is_default=True
                        )
                
                # Parse order date
                order_date_str = first_row["SIPARIŞ TARIHI VE SAATI"]
                try:
                    # Try to parse DD.MM.YYYY HH:MM format
                    order_date = pd.to_datetime(order_date_str, format="%d.%m.%Y %H:%M")
                except:
                    # If that fails, use pandas default parser with error handling
                    try:
                        order_date = pd.to_datetime(order_date_str)
                    except:
                        order_date = timezone.now()
                        result['error_rows'].append({
                            'row': row_num,
                            'error': f"Geçersiz tarih formatı: {order_date_str}, varsayılan tarih kullanıldı"
                        })
                
                # Create or update order
                if existing_order:
                    order = existing_order
                    # Update order fields if needed
                    order.customer = customer
                    order.order_date = order_date
                    # We're not updating status or payment status from the import
                else:
                    # Prepare order notes
                    location_info = []
                    if not pd.isna(first_row.get('EYALET', '')):
                        location_info.append(first_row.get('EYALET', ''))
                    if not pd.isna(first_row.get('ŞEHIR', '')):
                        location_info.append(first_row.get('ŞEHIR', ''))
                    
                    order_notes = "Imported from Excel"
                    if location_info:
                        order_notes += f": {', '.join(location_info)}"
                        
                    order = Order(
                        order_number=order_number,
                        customer=customer,
                        status='pending',  # Default to pending
                        payment_status='pending',  # Default to pending
                        order_date=order_date,
                        notes=order_notes,
                        owner=None  # Set owner to None explicitly
                    )
                
                # Save order
                order.save()
                
                # Clear existing order items if updating
                if existing_order and update_existing:
                    order.items.all().delete()
                
                # Process each row in this order group to create order items
                for idx, item_row in order_group.iterrows():
                    try:
                        item_row_num = idx + 2  # For error reporting
                        
                        # Get SKU and product name
                        sku = item_row["SKU"]
                        product_name = item_row["ÜRÜN ISMI"]
                        
                        if pd.isna(sku) or pd.isna(product_name):
                            raise ValueError("SKU ve ürün ismi zorunludur")
                        
                        # Parse quantity and price values
                        try:
                            quantity = int(item_row["ADET"])
                        except:
                            raise ValueError(f"Geçersiz adet: {item_row['ADET']}")
                        
                        # Handle comma in price (convert from European to US format)
                        unit_price_str = str(item_row["BIRIM FIYAT"]).replace(',', '.')
                        try:
                            # Convert directly to Decimal for better precision
                            try:
                                unit_price = Decimal(unit_price_str)
                            except InvalidOperation:
                                raise ValueError(f"Geçersiz birim fiyat: {item_row['BIRIM FIYAT']}")
                        except:
                            raise ValueError(f"Geçersiz birim fiyat: {item_row['BIRIM FIYAT']}")
                        
                        # Handle discount if present
                        discount_amount = 0
                        if "BIRIM INDIRIM" in item_row and not pd.isna(item_row["BIRIM INDIRIM"]):
                            discount_str = str(item_row["BIRIM INDIRIM"]).replace(',', '.')
                            try:
                                # Convert directly to Decimal for better precision
                                try:
                                    discount_amount = Decimal(discount_str)
                                except InvalidOperation:
                                    result['error_rows'].append({
                                        'row': item_row_num,
                                        'error': f"Geçersiz indirim tutarı: {item_row['BIRIM INDIRIM']}, 0 olarak alındı"
                                    })
                            except:
                                # If we can't parse discount, log an error but continue
                                result['error_rows'].append({
                                    'row': item_row_num,
                                    'error': f"Geçersiz indirim tutarı: {item_row['BIRIM INDIRIM']}, 0 olarak alındı"
                                })
                        
                        # Find or create product based on SKU or GTIN
                        product = None
                        
                        # First try to find by SKU
                        products = Product.objects.filter(sku=sku)
                        if products.exists():
                            product = products.first()
                        else:
                            # If not found by SKU, try to find by GTIN/barcode
                            if "GTIN" in item_row and not pd.isna(item_row.get("GTIN")):
                                gtin_raw = item_row.get("GTIN")
                                if isinstance(gtin_raw, float):
                                    gtin_value = str(int(gtin_raw))
                                else:
                                    gtin_value = str(gtin_raw)
                                
                                products_by_barcode = Product.objects.filter(barcode=gtin_value)
                                if products_by_barcode.exists():
                                    product = products_by_barcode.first()
                        
                        if not product:
                            # Get GTIN value and convert to string
                            gtin_value = ""
                            if "GTIN" in item_row and not pd.isna(item_row.get("GTIN")):
                                # Convert to string and remove decimal point if exists
                                gtin_raw = item_row.get("GTIN")
                                if isinstance(gtin_raw, float):
                                    gtin_value = str(int(gtin_raw))
                                else:
                                    gtin_value = str(gtin_raw)
                            
                            # Create a new product if not found
                            product = Product.objects.create(
                                name=product_name,
                                sku=sku,
                                barcode=gtin_value,
                                description="",  # Empty description
                                price=unit_price,
                                discount_price=0,
                                tax_rate=18,  # Default tax rate
                                stock=100,  # Default stock
                                is_active=True,
                                code=sku  # Ensure code is set (required field)
                            )
                        
                        # Ensure decimal fields have exactly 2 decimal places
                        try:
                            # Convert to Decimal and quantize to 2 decimal places
                            unit_price_decimal = Decimal(str(unit_price)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                            discount_amount_decimal = Decimal(str(discount_amount)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                            
                            # Create order item with properly formatted decimal values
                            OrderItem.objects.create(
                                order=order,
                                product=product,
                                quantity=quantity,
                                unit_price=unit_price_decimal,
                                discount_amount=discount_amount_decimal,
                                tax_rate=product.tax_rate
                            )
                        except Exception as e:
                            detailed_error = str(e)
                            # Log detailed error for debugging
                            logger.error(f"Row {item_row_num} error: {detailed_error}")
                            result['error_rows'].append({
                                'row': item_row_num,
                                'error': f"Sipariş kalemi hatası: {detailed_error}"
                            })
                            continue
                        
                    except Exception as e:
                        result['error_rows'].append({
                            'row': item_row_num,
                            'error': f"Sipariş kalemi hatası: {str(e)}"
                        })
                
                # Calculate order totals
                order.calculate_totals()
                
                if existing_order:
                    result['updated'] += 1
                else:
                    result['created'] += 1
                
            except Exception as e:
                result['error_count'] += 1
                result['error_rows'].append({
                    'row': row_num,
                    'error': f"Sipariş hatası: {str(e)}"
                })
        
        # Recalculate error_count based on the actual error rows
        result['error_count'] = len(result['error_rows'])
        
        return result
        
    except Exception as e:
        # Handle file reading/format errors
        raise Exception(f"Excel dosyası işlenirken hata oluştu: {str(e)}")


class OrderExcelImport:
    """Sipariş Excel import sınıfı - core.views_import tarafından kullanılıyor"""
    
    @staticmethod
    def import_data(file, user=None):
        """Excel dosyasından sipariş verilerini içe aktar"""
        return import_orders_excel(file)
    
    @staticmethod
    def get_template():
        """Sipariş import şablonunu oluştur"""
        return generate_order_template()